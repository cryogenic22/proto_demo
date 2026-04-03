"""
FastAPI backend for the Protocol Table Extraction Pipeline.

Provides endpoints for:
- PDF upload and extraction
- Status polling for long-running extractions
- Result retrieval
- Golden set evaluation
"""

from __future__ import annotations

import asyncio
import hashlib
import json
import logging
import os
import time
import uuid
from pathlib import Path
from typing import Any

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv optional in production

from fastapi import FastAPI, File, HTTPException, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse
from pydantic import BaseModel, Field

from src.models.schema import PipelineConfig, PipelineOutput
from src.persistence.ke_store import create_ke_store
from src.pipeline.orchestrator import PipelineOrchestrator

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI(
    title="Protocol Table Extractor",
    description="Extracts and digitizes tables from clinical trial protocol PDFs",
    version="0.1.0",
)

FRONTEND_URL = os.environ.get("FRONTEND_URL", "http://localhost:3000")

_allowed_origins = [
    "http://localhost:3000",
    "http://localhost:3001",
    FRONTEND_URL,
]


def _cors_origin_allowed(origin: str) -> bool:
    """Allow explicit origins + any *.up.railway.app subdomain."""
    if origin in _allowed_origins:
        return True
    if origin.endswith(".up.railway.app") and origin.startswith("https://"):
        return True
    return False


def _find_protocol_pdf(protocol_id: str) -> Path | None:
    """Find the PDF file for a protocol. Shared by page-image and verbatim endpoints."""
    import re as _re

    num_match = _re.match(r"^[pP][-_]?(\d+)", protocol_id)
    pid_dash = f"P-{num_match.group(1).zfill(2)}" if num_match else protocol_id
    pid_norm = protocol_id.upper().replace("_", "-")

    search_dirs = [Path("data/pdfs"), Path("golden_set/cached_pdfs")]

    for d in search_dirs:
        if not d.exists():
            continue
        for pattern in [f"{protocol_id}.pdf", f"{pid_dash}.pdf", f"{pid_norm}.pdf"]:
            candidate = d / pattern
            if candidate.exists():
                return candidate
        for p in d.glob("*.pdf"):
            stem_clean = _re.sub(r"[^a-zA-Z0-9]", "", p.stem).lower()
            pid_clean = _re.sub(r"[^a-zA-Z0-9]", "", protocol_id).lower()
            if stem_clean == pid_clean or stem_clean.startswith(pid_clean):
                return p

    # Fallback: look up document_name from stored protocol
    try:
        store = create_ke_store()
        protocol = store.load_protocol(protocol_id)
        if protocol:
            doc_name = getattr(protocol, "document_name", "") or ""
            if doc_name:
                for d in search_dirs:
                    if not d.exists():
                        continue
                    candidate = d / doc_name
                    if candidate.exists():
                        return candidate
                    for p in d.glob("*.pdf"):
                        if p.stem.lower().replace(" ", "_") == doc_name.replace(".pdf", "").lower().replace(" ", "_"):
                            return p
    except Exception:
        pass

    return None


app.add_middleware(
    CORSMiddleware,
    allow_origin_regex=r"https://.*\.up\.railway\.app",
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)


# Ensure CORS headers on ALL responses including errors.
# FastAPI's CORSMiddleware can miss HTTPException responses in some versions.
@app.exception_handler(HTTPException)
async def cors_http_exception_handler(request, exc):
    return JSONResponse(
        status_code=exc.status_code,
        content={"detail": exc.detail},
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "*",
            "Access-Control-Allow-Headers": "*",
        },
    )


@app.exception_handler(Exception)
async def cors_general_exception_handler(request, exc):
    return JSONResponse(
        status_code=500,
        content={"detail": str(exc)},
        headers={
            "Access-Control-Allow-Origin": "*",
            "Access-Control-Allow-Methods": "*",
            "Access-Control-Allow-Headers": "*",
        },
    )


# ---------------------------------------------------------------------------
# Persistent job store — survives server restarts
# ---------------------------------------------------------------------------

_JOBS_FILE = Path("data/jobs.json")


def _load_jobs() -> dict[str, dict[str, Any]]:
    """Load jobs from disk on startup."""
    if _JOBS_FILE.exists():
        try:
            return json.loads(_JOBS_FILE.read_text(encoding="utf-8"))
        except Exception:
            logger.warning("Could not load jobs.json, starting fresh")
    return {}


def _save_jobs() -> None:
    """Persist current jobs dict to disk."""
    try:
        _JOBS_FILE.parent.mkdir(parents=True, exist_ok=True)
        _JOBS_FILE.write_text(
            json.dumps(jobs, indent=2, default=str), encoding="utf-8"
        )
    except Exception as e:
        logger.warning("Could not save jobs.json: %s", e)


jobs: dict[str, dict[str, Any]] = _load_jobs()

# Mark any jobs that were "processing" when the server died as failed
for _jid, _jdata in jobs.items():
    if _jdata.get("status") == "processing":
        _jdata["status"] = "failed"
        _jdata["message"] = "Server restarted during extraction"
        _jdata["error"] = "Server restart — please re-upload"
_save_jobs()


@app.exception_handler(Exception)
async def global_exception_handler(request, exc):
    """Catch-all so the server never crashes on unhandled errors."""
    logger.exception(f"Unhandled error on {request.url}: {exc}")
    return JSONResponse(
        status_code=500,
        content={"detail": f"Internal server error: {str(exc)}"},
    )


class JobStatus(BaseModel):
    job_id: str
    status: str  # "pending", "processing", "completed", "failed"
    progress: int  # 0-100
    message: str
    document_name: str
    result: dict | None = None
    error: str | None = None
    created_at: float
    completed_at: float | None = None


# ---------------------------------------------------------------------------
# Protocol workspace request/response models
# ---------------------------------------------------------------------------


class AskRequest(BaseModel):
    """Request body for protocol Q&A."""

    question: str
    section_context: str | None = None


class AskSource(BaseModel):
    """Source citation in an assistant answer."""

    section: str = ""
    page: int = 0


class AskResponse(BaseModel):
    """Response from protocol Q&A."""

    role: str = "assistant"
    content: str = ""
    sources: list[AskSource] = Field(default_factory=list)


class ReviewRequest(BaseModel):
    """Request body for cell review actions."""

    table_id: str
    row: int
    col: int
    action: str  # "accept", "correct", "flag"
    correct_value: str | None = None
    flag_reason: str | None = None


class ReviewResponse(BaseModel):
    """Response from cell review."""

    success: bool = True


class ProcedureLibraryEntry(BaseModel):
    """A single procedure in the canonical library."""

    canonical_name: str
    cpt_code: str = ""
    code_system: str = ""
    category: str = ""
    cost_tier: str = ""
    aliases: str = ""


# Extraction speed presets
EXTRACTION_PRESETS = {
    "fast": {
        "max_extraction_passes": 1,      # Single pass (no dual-pass reconciliation)
        "enable_challenger": False,       # Skip adversarial validation
        "render_dpi": 120,               # Lower DPI for faster rendering
        "max_concurrent_llm_calls": 15,  # More parallel calls
    },
    "balanced": {
        "max_extraction_passes": 2,
        "enable_challenger": False,
        "render_dpi": 150,
        "max_concurrent_llm_calls": 10,
    },
    "thorough": {
        "max_extraction_passes": 2,
        "enable_challenger": True,
        "render_dpi": 150,
        "max_concurrent_llm_calls": 10,
    },
}

# Current preset — configurable via env or admin
_current_preset = os.environ.get("EXTRACTION_PRESET", "balanced")


def _build_config(**overrides) -> PipelineConfig:
    """Build PipelineConfig from environment variables + preset. Single source of truth."""
    preset = EXTRACTION_PRESETS.get(_current_preset, EXTRACTION_PRESETS["balanced"])
    merged = {**preset, **overrides}

    return PipelineConfig(
        llm_provider=os.environ.get("LLM_PROVIDER", "anthropic"),
        llm_model=os.environ.get("LLM_MODEL", ""),
        vision_model=os.environ.get("VISION_MODEL", ""),
        anthropic_api_key=os.environ.get("ANTHROPIC_API_KEY", ""),
        openai_api_key=os.environ.get("OPENAI_API_KEY", ""),
        azure_openai_api_key=os.environ.get("AZURE_OPENAI_API_KEY", ""),
        azure_openai_endpoint=os.environ.get("AZURE_OPENAI_ENDPOINT", ""),
        azure_openai_api_version=os.environ.get("AZURE_OPENAI_API_VERSION", "2024-12-01-preview"),
        azure_openai_deployment=os.environ.get("AZURE_OPENAI_DEPLOYMENT", ""),
        render_dpi=merged.get("render_dpi", 150),
        max_extraction_passes=merged.get("max_extraction_passes", 2),
        enable_challenger=merged.get("enable_challenger", True),
        max_concurrent_llm_calls=int(os.environ.get("MAX_CONCURRENT_LLM_CALLS", str(merged.get("max_concurrent_llm_calls", 10)))),
        openai_batch_mode=os.environ.get("OPENAI_BATCH_MODE", "").lower() in ("true", "1", "yes"),
        soa_only=True,
        **{k: v for k, v in overrides.items() if k not in merged},
    )


@app.get("/api/health")
async def health():
    return {"status": "ok", "version": "0.1.0"}


@app.get("/api/telemetry/runs")
async def get_telemetry_runs(limit: int = 20):
    """Get recent pipeline run telemetry."""
    from src.pipeline.telemetry import get_telemetry
    return {"runs": get_telemetry().get_recent_runs(limit)}


@app.get("/api/telemetry/errors")
async def get_telemetry_errors(limit: int = 50):
    """Get recent pipeline errors with tracebacks."""
    from src.pipeline.telemetry import get_telemetry
    return {"errors": get_telemetry().get_errors(limit)}


@app.get("/api/checkpoints/{doc_hash}")
async def get_checkpoint(doc_hash: str):
    """Recover extracted tables from a crashed pipeline run."""
    from src.pipeline.orchestrator import _CheckpointManager
    data = _CheckpointManager.load_checkpoint(doc_hash)
    if not data:
        raise HTTPException(status_code=404, detail="No checkpoint found")
    return data


@app.get("/api/benchmark")
async def get_benchmark():
    """Get benchmark comparison across all tested protocols."""
    from src.pipeline.benchmark import load_benchmark
    benchmarks = load_benchmark()
    return {"protocols": [b.__dict__ for b in benchmarks], "total": len(benchmarks)}


@app.get("/api/benchmark/report")
async def get_benchmark_report():
    """Get benchmark HTML report."""
    from fastapi.responses import HTMLResponse
    from src.pipeline.benchmark import generate_benchmark_html
    html = generate_benchmark_html()
    return HTMLResponse(html)


@app.get("/api/procedures/mapping")
async def get_procedure_mapping():
    """Export the full procedure mapping table for clinical review."""
    from src.pipeline.procedure_normalizer import ProcedureNormalizer
    normalizer = ProcedureNormalizer()
    return {"procedures": normalizer.get_mapping_table(), "total": len(normalizer.get_mapping_table())}


@app.post("/api/sections")
async def parse_sections(file: UploadFile = File(...), use_llm: bool = False):
    """Parse all sections from a PDF or DOCX. Returns the full document outline.

    Args:
        use_llm: Force LLM-assisted parsing (recommended for non-standard documents)
    """
    from src.pipeline.section_parser import SectionParser
    from src.llm.client import LLMClient

    file_bytes = await file.read()
    parser = SectionParser()
    sections = parser.parse(file_bytes, filename=file.filename or "")

    # Auto-trigger LLM fallback if deterministic parsing found too few sections
    if (use_llm or parser.needs_llm_fallback) and len(sections) < 10:
        logger.info(f"Section parser found {len(sections)} sections — using LLM fallback")
        config = _build_config()
        llm = LLMClient(config)
        llm_sections = await parser.parse_with_llm(file_bytes, llm_client=llm)
        if len(llm_sections) > len(sections):
            sections = llm_sections

    return {
        "sections": [s.to_dict() for s in sections],
        "total": len(sections),
        "outline": parser.to_outline(sections),
        "method": "llm_fallback" if parser.needs_llm_fallback else "deterministic",
    }


@app.post("/api/verbatim")
async def extract_verbatim(
    file: UploadFile = File(...),
    instruction: str = "",
    include_subsections: bool = True,
    output_format: str = "text",
):
    """Extract verbatim content from a protocol PDF/DOCX.

    The LLM locates the content; PyMuPDF/python-docx extracts the exact text.
    Zero hallucination — the output text is never LLM-generated.

    Args:
        instruction: What to extract (e.g., "Copy Section 5.1")
        include_subsections: Include subsection content (default True)
        output_format: "text" (plain), "html" (semantic HTML with paragraphs, lists, tables),
            "docx" (Word document download)

    Example instructions:
    - "Copy Section 5.1"
    - "Extract the inclusion criteria"
    - "Get the Schedule of Activities table"
    - "Copy the primary endpoint definition from Section 3"
    """
    if not instruction:
        raise HTTPException(status_code=400, detail="Provide an 'instruction' query parameter")

    from src.pipeline.verbatim_extractor import VerbatimExtractor
    file_bytes = await file.read()
    config = _build_config()
    extractor = VerbatimExtractor(config)
    result = await extractor.extract(
        file_bytes, instruction, filename=file.filename or "",
        output_format=output_format,
    )

    # DOCX output: return as downloadable file
    if output_format == "docx" and result.sections_found:
        from fastapi.responses import Response
        section = extractor.section_parser.find(
            extractor.section_parser.parse(file_bytes, filename=file.filename or ""),
            result.sections_found[0],
        )
        if section:
            docx_bytes = extractor.section_parser.get_section_formatted(
                file_bytes, section, output="docx"
            )
            return Response(
                content=docx_bytes,
                media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
                headers={"Content-Disposition": f'attachment; filename="section_{result.sections_found[0]}.docx"'},
            )

    return {
        "instruction": result.instruction,
        "sections_found": result.sections_found,
        "content_type": result.content_type,
        "text": result.text,
        "tables": result.tables,
        "source_pages": result.source_pages,
        "explanation": result.explanation,
        "is_verbatim": result.is_verbatim,
    }


@app.post("/api/protocols/{protocol_id}/verbatim")
async def extract_verbatim_from_protocol(
    protocol_id: str,
    body: dict,
):
    """Extract verbatim content from a stored protocol's PDF.

    No file upload needed — uses the PDF stored on the server.
    Body: {"instruction": "Copy Section 5.1", "output_format": "html"}
    """
    instruction = body.get("instruction", "")
    output_format = body.get("output_format", "html")

    if not instruction:
        raise HTTPException(status_code=400, detail="Provide an 'instruction'")

    pdf_path = _find_protocol_pdf(protocol_id)
    if not pdf_path:
        raise HTTPException(
            status_code=404,
            detail=f"PDF not found for {protocol_id}. Upload the document to use verbatim extraction.",
        )

    from src.pipeline.verbatim_extractor import VerbatimExtractor

    file_bytes = pdf_path.read_bytes()
    config = _build_config()
    extractor = VerbatimExtractor(config)
    result = await extractor.extract(
        file_bytes, instruction, filename=pdf_path.name,
        output_format=output_format,
    )

    return {
        "instruction": result.instruction,
        "sections_found": result.sections_found,
        "content_type": result.content_type,
        "text": result.text,
        "tables": result.tables,
        "source_pages": result.source_pages,
        "explanation": result.explanation,
        "is_verbatim": result.is_verbatim,
    }


@app.post("/api/procedures/check")
async def check_procedure_mapping(names: list[str]):
    """Check which procedure names can/cannot be mapped."""
    from src.pipeline.procedure_normalizer import ProcedureNormalizer
    normalizer = ProcedureNormalizer()
    mapped = [normalizer.normalize(n).model_dump() for n in names]
    unmapped = normalizer.get_unmapped_report(names)
    return {"mapped": mapped, "unmapped": unmapped}


@app.post("/api/extract")
async def extract_protocol(file: UploadFile = File(...)):
    """
    Upload a protocol PDF and start extraction.
    Returns a job_id for polling status.
    """
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(status_code=400, detail="Only PDF files are accepted")

    pdf_bytes = await file.read()
    if len(pdf_bytes) == 0:
        raise HTTPException(status_code=400, detail="Empty file")

    if len(pdf_bytes) > 100 * 1024 * 1024:  # 100MB limit
        raise HTTPException(status_code=400, detail="File too large (max 100MB)")

    job_id = uuid.uuid4().hex[:12]
    jobs[job_id] = {
        "status": "pending",
        "progress": 0,
        "message": "Queued for processing",
        "document_name": file.filename,
        "result": None,
        "error": None,
        "created_at": time.time(),
        "completed_at": None,
    }

    _save_jobs()

    # Run extraction in background with error logging
    task = asyncio.create_task(_run_extraction(job_id, pdf_bytes, file.filename))
    task.add_done_callback(_task_done_callback)

    return {"job_id": job_id, "status": "pending"}


def _task_done_callback(task: asyncio.Task):
    """Log any unhandled exception from background tasks so they don't crash the server."""
    if task.cancelled():
        logger.warning("Extraction task was cancelled")
        return
    exc = task.exception()
    if exc:
        logger.error(f"Extraction task failed with unhandled exception: {exc}", exc_info=exc)


@app.get("/api/jobs/{job_id}")
async def get_job_status(job_id: str):
    """Get the status of an extraction job."""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")

    job = jobs[job_id]
    return JobStatus(
        job_id=job_id,
        status=job["status"],
        progress=job["progress"],
        message=job["message"],
        document_name=job["document_name"],
        result=job["result"],
        error=job["error"],
        created_at=job["created_at"],
        completed_at=job["completed_at"],
    )


@app.get("/api/jobs")
async def list_jobs():
    """List all jobs."""
    return [
        JobStatus(
            job_id=jid,
            status=j["status"],
            progress=j["progress"],
            message=j["message"],
            document_name=j["document_name"],
            result=None,  # Don't include full results in list view
            error=j["error"],
            created_at=j["created_at"],
            completed_at=j["completed_at"],
        )
        for jid, j in sorted(jobs.items(), key=lambda x: x[1]["created_at"], reverse=True)
    ]


@app.delete("/api/jobs")
async def clear_all_jobs():
    """Clear all jobs from the queue."""
    count = len(jobs)
    jobs.clear()
    _save_jobs()
    return {"cleared": count}


@app.delete("/api/jobs/{job_id}")
async def delete_job(job_id: str):
    """Delete a specific job."""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")
    del jobs[job_id]
    _save_jobs()
    return {"deleted": job_id}


# ── Admin Endpoints ──────────────────────────────────────────────────────

@app.delete("/api/admin/protocols/{protocol_id}")
async def admin_delete_protocol(protocol_id: str):
    """Delete a stored protocol and ALL related data."""
    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if not protocol:
        raise HTTPException(status_code=404, detail=f"Protocol {protocol_id} not found")

    deleted_files = []

    # Delete protocol JSON
    proto_path = Path(f"data/protocols/{protocol_id}.json")
    if proto_path.exists():
        proto_path.unlink()
        deleted_files.append(str(proto_path))

    # Delete KE sidecar
    ke_path = Path(f"data/protocols/{protocol_id}_kes.json")
    if ke_path.exists():
        ke_path.unlink()
        deleted_files.append(str(ke_path))

    # Delete annotations
    for ann_dir in [Path("data/annotations"), Path("golden_set/annotations")]:
        if ann_dir.exists():
            for f in ann_dir.glob(f"*{protocol_id}*"):
                f.unlink()
                deleted_files.append(str(f))

    # Delete related jobs
    related_jobs = [
        jid for jid, j in jobs.items()
        if j.get("protocol_id") == protocol_id
        or protocol_id in j.get("document_name", "")
    ]
    for jid in related_jobs:
        del jobs[jid]
    if related_jobs:
        _save_jobs()

    # Clear caches
    _smb_cache.pop(protocol_id, None)
    _trust_cache.pop(protocol_id, None)

    return {
        "deleted": protocol_id,
        "files_removed": deleted_files,
        "jobs_removed": len(related_jobs),
    }


@app.delete("/api/admin/jobs")
async def admin_clear_all_jobs():
    """Clear all extraction jobs."""
    count = len(jobs)
    jobs.clear()
    _save_jobs()
    return {"cleared": count}


@app.get("/api/admin/config")
async def admin_get_config():
    """Get current extraction configuration."""
    return {
        "current_preset": _current_preset,
        "presets": EXTRACTION_PRESETS,
        "config": _build_config().model_dump(),
    }


@app.put("/api/admin/config/preset/{preset_name}")
async def admin_set_preset(preset_name: str):
    """Switch extraction speed preset: fast, balanced, thorough."""
    global _current_preset
    if preset_name not in EXTRACTION_PRESETS:
        raise HTTPException(
            status_code=400,
            detail=f"Unknown preset '{preset_name}'. Available: {list(EXTRACTION_PRESETS.keys())}",
        )
    _current_preset = preset_name
    return {"preset": preset_name, "config": EXTRACTION_PRESETS[preset_name]}


@app.get("/api/admin/stats")
async def admin_stats():
    """Get system statistics for admin dashboard."""
    store = create_ke_store()
    protocols = store.list_protocols()
    proto_list = []
    for pid in protocols:
        p = store.load_protocol(pid)
        if p:
            pdata = p.model_dump(mode="json") if hasattr(p, "model_dump") else p
            meta = pdata.get("metadata", {})
            tables = pdata.get("tables", [])
            proto_list.append({
                "protocol_id": pid,
                "document_name": pdata.get("document_name", ""),
                "title": meta.get("title", "") if isinstance(meta, dict) else "",
                "tables_count": len(tables),
                "total_cells": sum(len(t.get("cells", [])) for t in tables),
                "total_procedures": sum(len(t.get("procedures", [])) for t in tables),
            })

    job_list = []
    for jid, jdata in sorted(jobs.items(), key=lambda x: x[1].get("created_at", 0), reverse=True):
        job_list.append({
            "job_id": jid,
            "status": jdata.get("status", "unknown"),
            "document_name": jdata.get("document_name", ""),
            "progress": jdata.get("progress", 0),
            "message": jdata.get("message", ""),
            "created_at": jdata.get("created_at"),
        })

    return {
        "protocols": proto_list,
        "jobs": job_list,
        "total_protocols": len(proto_list),
        "total_jobs": len(job_list),
    }


@app.get("/api/jobs/{job_id}/result")
async def get_job_result(job_id: str):
    """Get the full result of a completed extraction job."""
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")

    job = jobs[job_id]
    if job["status"] != "completed":
        raise HTTPException(status_code=400, detail=f"Job is {job['status']}, not completed")

    return job["result"]


@app.get("/api/jobs/{job_id}/review")
async def get_job_review(job_id: str, format: str = "json"):
    """Get extraction results formatted for medical writer review.

    Args:
        format: "json" for structured review, "markdown" for human-readable doc
    """
    if job_id not in jobs:
        raise HTTPException(status_code=404, detail="Job not found")

    job = jobs[job_id]
    if job["status"] != "completed":
        raise HTTPException(status_code=400, detail=f"Job is {job['status']}, not completed")

    from src.models.schema import PipelineOutput as PipelineOutputModel
    from src.pipeline.review_exporter import export_review_document, export_review_json

    result = PipelineOutputModel.model_validate(job["result"])

    if format == "markdown":
        from fastapi.responses import PlainTextResponse
        md = export_review_document(result)
        return PlainTextResponse(md, media_type="text/markdown")
    elif format == "html":
        from fastapi.responses import HTMLResponse
        from src.pipeline.html_report import generate_html_report
        from src.pipeline.run_comparator import compare_runs
        # Compare against previous run if available
        comparison = compare_runs(job["result"])
        comparison_html = comparison.to_html_section() if comparison else ""
        html = generate_html_report(result, comparison_html=comparison_html)
        return HTMLResponse(html)
    elif format == "budget":
        from fastapi.responses import HTMLResponse
        from src.pipeline.budget_calculator import generate_budget_html
        html = generate_budget_html(result)
        return HTMLResponse(html)
    else:
        return export_review_json(result)


async def _run_extraction(job_id: str, pdf_bytes: bytes, filename: str):
    """Background task to run the extraction pipeline.

    Wrapped in broad exception handling so the server NEVER crashes —
    any failure is captured and reported back through the job status.
    """
    from src.pipeline.telemetry import get_telemetry
    tel = get_telemetry()
    start_time = time.time()

    try:
        jobs[job_id]["status"] = "processing"
        jobs[job_id]["progress"] = 5
        jobs[job_id]["message"] = "Initializing pipeline..."

        config = _build_config()
        orchestrator = PipelineOrchestrator(config)

        jobs[job_id]["progress"] = 10
        jobs[job_id]["message"] = "Ingesting PDF..."

        tel.log_run_start(job_id, filename, 0, config.model_dump())

        def on_progress(pct: int, msg: str):
            jobs[job_id]["progress"] = pct
            jobs[job_id]["message"] = msg
            tel.log_stage(job_id, "progress", detail=f"{pct}% {msg}")

        result = await orchestrator.run(pdf_bytes, filename, on_progress=on_progress)

        # Serialize result — use try/except to handle serialization errors
        try:
            result_json = json.loads(result.model_dump_json())
        except Exception as ser_err:
            logger.error(f"Result serialization failed: {ser_err}")
            result_json = {
                "document_name": filename,
                "tables": [],
                "total_pages": result.total_pages,
                "warnings": result.warnings + [f"Serialization error: {ser_err}"],
            }

        total_cells = sum(len(t.cells) for t in result.tables)
        total_fn = sum(len(t.footnotes) for t in result.tables)
        avg_conf = sum(t.overall_confidence for t in result.tables) / max(len(result.tables), 1)

        jobs[job_id]["status"] = "completed"
        jobs[job_id]["progress"] = 100
        jobs[job_id]["message"] = f"Extracted {len(result.tables)} tables"
        jobs[job_id]["result"] = result_json
        jobs[job_id]["completed_at"] = time.time()
        _save_jobs()

        # Persist protocol to the knowledge-element store
        try:
            from src.persistence.protocol_bridge import (
                pipeline_output_to_protocol,
            )

            protocol = pipeline_output_to_protocol(result_json, filename, pdf_bytes=pdf_bytes)
            store = create_ke_store()
            store.save_protocol(protocol)
            logger.info("Saved protocol %s to store", protocol.protocol_id)
        except Exception as bridge_err:
            logger.warning("Protocol persistence failed: %s", bridge_err)

        # Auto-record benchmark
        try:
            from src.pipeline.benchmark import from_pipeline_output, add_benchmark
            bm = from_pipeline_output(result_json, job_id)
            add_benchmark(bm)
        except Exception as bm_err:
            logger.warning(f"Benchmark recording failed: {bm_err}")

        tel.log_run_end(
            job_id, "completed", tables=len(result.tables), cells=total_cells,
            footnotes=total_fn, confidence=avg_conf,
            duration_s=time.time() - start_time, warnings=len(result.warnings),
        )

    except MemoryError:
        logger.error(f"Out of memory processing job {job_id}")
        jobs[job_id]["status"] = "failed"
        jobs[job_id]["progress"] = 0
        jobs[job_id]["message"] = "Out of memory — try a smaller document or reduce DPI"
        jobs[job_id]["error"] = "MemoryError: document too large"
        jobs[job_id]["completed_at"] = time.time()
        tel.log_run_end(job_id, "failed_oom", 0, 0, 0, 0, time.time() - start_time,
                        errors=["MemoryError"])

    except Exception as e:
        logger.exception(f"Extraction failed for job {job_id}")
        jobs[job_id]["status"] = "failed"
        jobs[job_id]["progress"] = 0
        jobs[job_id]["message"] = f"Extraction failed: {str(e)}"
        jobs[job_id]["error"] = str(e)
        jobs[job_id]["completed_at"] = time.time()
        import traceback
        tel.log_run_end(job_id, "failed", 0, 0, 0, 0, time.time() - start_time,
                        errors=[str(e)])
        tel.log_error(job_id, "pipeline", str(e), traceback.format_exc())

    finally:
        _save_jobs()
        import gc
        gc.collect()


# ---------------------------------------------------------------------------
# Protocol workspace endpoints (ask, review, procedure library)
# ---------------------------------------------------------------------------


def _format_kg_context_for_agent(protocol_id: str) -> str:
    """Format cached SMB model data as text context for the LLM agent."""
    if protocol_id not in _smb_cache:
        return "(Knowledge graph not built for this protocol.)"

    cached = _smb_cache[protocol_id]
    graph = cached.get("graph", {})
    timeline = cached.get("timeline", [])
    schedule = cached.get("schedule", [])
    rules = cached.get("inference_rules_fired", [])

    lines: list[str] = []

    # ── Visits ──
    if timeline:
        lines.append("VISITS:")
        for v in timeline:
            day = f"Day {v['day_number']}" if v.get("day_number") is not None else "No day"
            window = ""
            if v.get("window_minus") or v.get("window_plus"):
                window = f" (window: -{v.get('window_minus', 0)}/+{v.get('window_plus', 0)} days)"
            lines.append(f"  - {v['visit_name']}: {day}, {v.get('procedure_count', 0)} procedures{window}")

    # ── Procedures & schedule ──
    if schedule:
        lines.append("\nPROCEDURES (schedule summary):")
        for s in schedule:
            name = s.get("canonical_name") or s.get("procedure", "Unknown")
            firm = s.get("firm_occurrences", 0)
            cond = s.get("conditional_occurrences", 0)
            cpt = s.get("cpt_code", "")
            visits = s.get("visits_required", [])
            visit_str = f" at visits: {', '.join(visits[:8])}" if visits else ""
            cpt_str = f" [CPT {cpt}]" if cpt else ""
            lines.append(f"  - {name}: {firm} firm, {cond} conditional{cpt_str}{visit_str}")

    # ── Footnotes ──
    footnotes = [n for n in graph.get("nodes", []) if n.get("type") == "Footnote"]
    if footnotes:
        lines.append("\nFOOTNOTES:")
        for fn in footnotes:
            marker = fn.get("properties", {}).get("footnote_marker", "?")
            text = fn.get("properties", {}).get("footnote_text", fn.get("label", ""))
            classification = fn.get("properties", {}).get("classification", "")
            cls_str = f" [{classification}]" if classification else ""
            lines.append(f"  {marker}: {text}{cls_str}")

    # ── Inference rules ──
    if rules:
        lines.append(f"\nINFERENCE RULES APPLIED: {', '.join(rules)}")

    return "\n".join(lines) if lines else "(No KG data available.)"


def _get_relevant_sections(protocol: Any, question: str) -> str:
    """Extract relevant section text from the protocol based on question keywords."""
    from src.models.protocol import SectionNode

    all_sections: list[SectionNode] = protocol._flatten_sections(protocol.sections)
    if not all_sections:
        return "(No parsed sections available.)"

    # Tokenize question into keywords (lowercase, skip short words)
    stop_words = {"the", "a", "an", "is", "are", "was", "were", "what", "how",
                  "many", "much", "does", "this", "that", "for", "and", "or",
                  "in", "of", "to", "with", "on", "at", "by", "from", "do"}
    keywords = [
        w.lower().strip("?.,!;:")
        for w in question.split()
        if len(w) > 2 and w.lower().strip("?.,!;:") not in stop_words
    ]

    # Score each section by keyword overlap in title + content
    scored: list[tuple[float, SectionNode]] = []
    for sec in all_sections:
        title_lower = (sec.title or "").lower()
        content_lower = (sec.content_html or "").lower()
        combined = f"{title_lower} {content_lower}"
        score = sum(1 for kw in keywords if kw in combined)
        # Boost sections with direct title match
        if any(kw in title_lower for kw in keywords):
            score += 2
        if score > 0:
            scored.append((score, sec))

    scored.sort(key=lambda x: -x[0])

    # Take top 5 relevant sections, limit total text
    result_parts: list[str] = []
    total_chars = 0
    max_chars = 8000  # Keep context manageable for the LLM
    for _score, sec in scored[:5]:
        header = f"[Section {sec.number}: {sec.title} (p.{sec.page})]"
        content = sec.content_html or "(no content)"
        # Strip HTML tags for cleaner LLM input
        import re
        content = re.sub(r"<[^>]+>", " ", content)
        content = re.sub(r"\s+", " ", content).strip()
        if total_chars + len(content) > max_chars:
            content = content[:max_chars - total_chars] + "..."
        result_parts.append(f"{header}\n{content}")
        total_chars += len(content)
        if total_chars >= max_chars:
            break

    if not result_parts:
        # Fallback: return first few sections
        for sec in all_sections[:3]:
            header = f"[Section {sec.number}: {sec.title} (p.{sec.page})]"
            content = sec.content_html or "(no content)"
            import re
            content = re.sub(r"<[^>]+>", " ", content)
            content = re.sub(r"\s+", " ", content).strip()[:2000]
            result_parts.append(f"{header}\n{content}")

    return "\n\n".join(result_parts) if result_parts else "(No section text available.)"


@app.post("/api/protocols/{protocol_id}/ask", response_model=AskResponse)
async def ask_protocol_endpoint(protocol_id: str, body: AskRequest):
    """Ask a question about a stored protocol — grounded in KG + section text."""
    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if protocol is None:
        raise HTTPException(status_code=404, detail="Protocol not found")

    # Build SMB model if not already cached
    if protocol_id not in _smb_cache:
        try:
            from src.smb.core.engine import SMBEngine
            from src.smb.core.query import get_budget_schedule, get_visit_timeline

            protocol_data = (
                protocol.model_dump(mode="json")
                if hasattr(protocol, "model_dump")
                else protocol
            )
            engine = SMBEngine(domain="protocol")
            result = engine.build_from_protocol_json(protocol_data)

            _smb_cache[protocol_id] = {
                "model": result.model.model_dump(mode="json"),
                "graph": result.model.to_graph_dict(),
                "summary": result.model.summary(),
                "schedule": get_budget_schedule(result.model),
                "timeline": get_visit_timeline(result.model),
                "build_time_seconds": result.build_time_seconds,
                "inference_rules_fired": result.inference_rules_fired,
                "validation_passed": result.validation_passed,
                "validation_errors": result.validation_errors,
                "validation_warnings": result.validation_warnings,
            }
        except Exception as e:
            logger.warning(f"Could not build SMB for ask endpoint: {e}")

    # Format KG context
    kg_context = _format_kg_context_for_agent(protocol_id)

    # Get relevant section text
    sections_text = _get_relevant_sections(protocol, body.question)

    # Collect source pages from relevant sections
    source_sections: list[AskSource] = []
    for sec in protocol._flatten_sections(protocol.sections):
        title_lower = (sec.title or "").lower()
        q_lower = body.question.lower()
        if any(w in title_lower for w in q_lower.split() if len(w) > 3):
            source_sections.append(AskSource(section=f"{sec.number} {sec.title}", page=sec.page))
    source_sections = source_sections[:5]  # Limit sources

    # Send to LLM
    try:
        from src.llm.client import LLMClient

        config = _build_config()
        llm = LLMClient(config)
        system_prompt = (
            "You are a clinical protocol analyst. Answer the question based ONLY on the "
            "provided context below. Be specific — cite visit names, procedure names, "
            "day numbers, and footnote markers when relevant. If the answer is not in "
            "the context, say so clearly.\n\n"
            f"KNOWLEDGE GRAPH:\n{kg_context}\n\n"
            f"PROTOCOL SECTIONS:\n{sections_text}"
        )
        response_text = await llm.text_query(
            prompt=body.question,
            system=system_prompt,
            max_tokens=2048,
            temperature=0.1,
        )
        return AskResponse(
            role="assistant",
            content=response_text,
            sources=source_sections if source_sections else [AskSource(section="knowledge_graph", page=0)],
        )
    except Exception as e:
        logger.exception(f"LLM query failed for protocol {protocol_id}: {e}")
        # Graceful fallback — return KG summary instead of an error
        fallback = (
            f"I couldn't reach the LLM service, but here is what the knowledge graph contains "
            f"for protocol '{protocol.document_name}':\n\n{kg_context}"
        )
        return AskResponse(
            role="assistant",
            content=fallback,
            sources=[AskSource(section="knowledge_graph", page=0)],
        )


@app.post("/api/protocols/{protocol_id}/review", response_model=ReviewResponse)
async def review_protocol_cell(protocol_id: str, body: ReviewRequest):
    """Accept, correct, or flag a cell in a stored protocol."""
    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if protocol is None:
        raise HTTPException(status_code=404, detail="Protocol not found")

    table = next(
        (t for t in protocol.tables if t.get("table_id") == body.table_id),
        None,
    )
    if table is None:
        raise HTTPException(status_code=404, detail="Table not found")

    cells = table.get("cells", [])
    cell = next(
        (c for c in cells if c.get("row") == body.row and c.get("col") == body.col),
        None,
    )

    old_value = cell.get("raw_value", "") if cell else ""
    old_confidence = cell.get("confidence", 0) if cell else 0

    if body.action == "accept" and cell is not None:
        cell["confidence"] = 1.0
        cell["human_reviewed"] = True
    elif body.action == "correct" and cell is not None:
        cell["raw_value"] = body.correct_value or ""
        cell["confidence"] = 1.0
        cell["human_reviewed"] = True
        cell["original_value"] = old_value
    elif body.action == "flag":
        review_items = table.setdefault("review_items", [])
        review_items.append({
            "row": body.row,
            "col": body.col,
            "reason": body.flag_reason or "",
            "action": "flag",
        })
        if cell is not None:
            cell["human_reviewed"] = True
    else:
        raise HTTPException(
            status_code=400, detail="Invalid action or cell not found"
        )

    store.save_protocol(protocol)

    # Append to ground truth annotations log
    _log_annotation(
        protocol_id=protocol_id,
        table_id=body.table_id,
        row=body.row,
        col=body.col,
        action=body.action,
        old_value=old_value,
        new_value=body.correct_value if body.action == "correct" else old_value,
        old_confidence=old_confidence,
        row_header=cell.get("row_header", "") if cell else "",
        col_header=cell.get("col_header", "") if cell else "",
    )

    # Invalidate trust cache for this protocol
    _trust_cache.pop(protocol_id, None)

    return ReviewResponse(success=True)


def _log_annotation(
    protocol_id: str,
    table_id: str,
    row: int,
    col: int,
    action: str,
    old_value: str,
    new_value: str,
    old_confidence: float,
    row_header: str = "",
    col_header: str = "",
) -> None:
    """Append human review annotation to ground truth log.

    Each annotation enriches the ground truth for future pipeline
    evaluation. Corrections become the authoritative cell value.
    """
    import csv
    from datetime import datetime, timezone

    gt_dir = Path("data/annotations")
    gt_dir.mkdir(parents=True, exist_ok=True)
    gt_file = gt_dir / f"{protocol_id}_annotations.csv"

    is_new = not gt_file.exists()
    with open(gt_file, "a", newline="", encoding="utf-8") as f:
        writer = csv.writer(f)
        if is_new:
            writer.writerow([
                "timestamp", "protocol_id", "table_id",
                "row", "col", "row_header", "col_header",
                "action", "old_value", "new_value",
                "old_confidence", "new_confidence",
            ])
        writer.writerow([
            datetime.now(timezone.utc).isoformat(),
            protocol_id, table_id,
            row, col, row_header, col_header,
            action, old_value, new_value,
            f"{old_confidence:.4f}", "1.0000",
        ])
    logger.info(
        f"Annotation logged: {protocol_id}/{table_id} "
        f"({row},{col}) {action}"
    )


@app.get("/api/procedures/library")
async def get_procedures_library(category: str = "", q: str = ""):
    """Return the procedure library with optional filtering."""
    from src.domain.vocabulary import get_procedure_vocab

    vocab = get_procedure_vocab()

    if q.strip():
        entries = vocab.search(q.strip())
    elif category.strip():
        entries = vocab.list_by_category(category.strip())
    else:
        entries = vocab.list_all()

    return [
        {
            "canonical_name": e.canonical_name,
            "cpt_code": e.cpt_code,
            "code_system": e.code_system,
            "category": e.category,
            "cost_tier": e.cost_tier,
            "aliases": e.aliases,
            "used_in_protocols": e.used_in_protocols,
        }
        for e in entries
    ]


@app.get("/api/procedures/library/search")
async def search_procedures(q: str = "", limit: int = 20):
    """Fuzzy search the procedure library by name, alias, or CPT code."""
    from src.domain.vocabulary import get_procedure_vocab

    vocab = get_procedure_vocab()
    if not q.strip():
        entries = vocab.list_all()[:limit]
    else:
        entries = vocab.search(q.strip())[:limit]

    return [e.to_dict() for e in entries]


@app.get("/api/procedures/library/hierarchies")
async def get_procedure_hierarchies():
    """Return procedure family hierarchies (parent-child relationships)."""
    from src.domain.vocabulary import get_procedure_hierarchy

    mgr = get_procedure_hierarchy()
    return [f.to_dict() for f in mgr.list_families()]


@app.get("/api/procedures/library/stats")
async def get_procedure_stats():
    """Return procedure library statistics."""
    from src.domain.vocabulary import get_procedure_vocab
    vocab = get_procedure_vocab()
    return vocab.get_stats()


@app.post("/api/procedures/library/import")
async def import_procedures_csv(file: UploadFile = File(...)):
    """Import procedures from a CSV file.

    Expected CSV format (header row required):
    canonical_name,cpt_code,code_system,category,cost_tier,aliases

    Existing procedures are updated, new ones are added.
    """
    from src.domain.vocabulary import get_procedure_vocab
    from src.domain.vocabulary.procedure_vocab import ProcedureEntry
    import csv
    from io import StringIO

    content = (await file.read()).decode("utf-8")
    reader = csv.DictReader(StringIO(content))

    vocab = get_procedure_vocab()
    added = 0
    updated = 0

    for row in reader:
        name = (row.get("canonical_name") or "").strip()
        if not name:
            continue

        existing = vocab.lookup(name)
        if existing:
            # Update fields if provided
            if row.get("cpt_code"):
                existing.cpt_code = row["cpt_code"].strip()
            if row.get("category"):
                existing.category = row["category"].strip()
            if row.get("cost_tier"):
                existing.cost_tier = row["cost_tier"].strip()
            if row.get("aliases"):
                new_aliases = [
                    a.strip() for a in row["aliases"].split(",") if a.strip()
                ]
                for alias in new_aliases:
                    if alias.lower() not in [
                        a.lower() for a in existing.aliases
                    ]:
                        existing.aliases.append(alias)
            updated += 1
        else:
            aliases = []
            if row.get("aliases"):
                aliases = [
                    a.strip() for a in row["aliases"].split(",") if a.strip()
                ]
            entry = ProcedureEntry(
                canonical_name=name,
                cpt_code=(row.get("cpt_code") or "").strip(),
                code_system=(row.get("code_system") or "CPT").strip(),
                category=(row.get("category") or "Unknown").strip(),
                cost_tier=(row.get("cost_tier") or "LOW").strip(),
                aliases=aliases,
            )
            vocab.add_procedure(entry)
            added += 1

    vocab._save()
    return {
        "added": added,
        "updated": updated,
        "total": len(vocab.list_all()),
    }


@app.get("/api/procedures/library/export")
async def export_procedures_csv():
    """Export the full procedure library as CSV."""
    from fastapi.responses import Response
    from src.domain.vocabulary import get_procedure_vocab

    vocab = get_procedure_vocab()
    lines = ["canonical_name,cpt_code,code_system,category,cost_tier,aliases"]
    for entry in vocab.list_all():
        aliases = ", ".join(entry.aliases)
        lines.append(
            f'"{entry.canonical_name}","{entry.cpt_code}",'
            f'"{entry.code_system}","{entry.category}",'
            f'"{entry.cost_tier}","{aliases}"'
        )

    return Response(
        content="\n".join(lines),
        media_type="text/csv",
        headers={
            "Content-Disposition": 'attachment; filename="procedure_library.csv"'
        },
    )


@app.get("/api/protocols/{protocol_id}/budget/lines")
async def get_budget_lines(protocol_id: str):
    """Get budget lines re-normalized with the current procedure vocabulary.

    This ensures CPT codes and categories reflect the latest vocabulary,
    not the stale values from extraction time.
    """
    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if not protocol:
        raise HTTPException(status_code=404, detail="Protocol not found")

    from src.pipeline.procedure_normalizer import ProcedureNormalizer

    normalizer = ProcedureNormalizer()
    COST_MAP = {"LOW": 75, "MEDIUM": 350, "HIGH": 1200, "VERY_HIGH": 3500}

    # If protocol has pre-computed budget lines, re-normalize them
    if protocol.budget_lines:
        lines = []
        for bl in protocol.budget_lines:
            if isinstance(bl, dict):
                raw = bl.get("procedure", bl.get("canonical_name", ""))
            else:
                raw = getattr(bl, "procedure", getattr(bl, "canonical_name", ""))

            # Skip non-procedures
            if normalizer.is_not_procedure(raw):
                continue

            # Re-normalize with current vocabulary
            norm = normalizer.normalize(raw)

            if isinstance(bl, dict):
                bl_out = dict(bl)
            else:
                bl_out = bl.__dict__ if hasattr(bl, "__dict__") else {}

            bl_out["cpt_code"] = norm.code or bl_out.get("cpt_code", "")
            bl_out["canonical_name"] = norm.canonical_name
            bl_out["category"] = norm.category if norm.category != "Unknown" else bl_out.get("category", "Unknown")
            bl_out["cost_tier"] = norm.estimated_cost_tier.value if norm.category != "Unknown" else bl_out.get("cost_tier", "LOW")
            bl_out["estimated_unit_cost"] = COST_MAP.get(bl_out["cost_tier"], bl_out.get("estimated_unit_cost", 75))
            lines.append(bl_out)
        return lines

    # Generate from table data if no pre-computed lines
    lines = []
    seen: set[str] = set()
    for table in protocol.tables:
        if not isinstance(table, dict):
            continue
        for proc in table.get("procedures", []):
            raw = proc.get("raw_name", "")
            if not raw or normalizer.is_not_procedure(raw):
                continue

            norm = normalizer.normalize(raw)
            key = norm.canonical_name.lower()
            if key in seen:
                continue
            seen.add(key)

            # Count marker cells
            markers = sum(
                1 for c in table.get("cells", [])
                if c.get("row_header", "").lower()[:20] == raw.lower()[:20]
                and c.get("data_type") == "MARKER"
            )

            tier = norm.estimated_cost_tier.value
            lines.append({
                "procedure": raw,
                "canonical_name": norm.canonical_name,
                "cpt_code": norm.code or "",
                "category": norm.category,
                "cost_tier": tier,
                "visits_required": [],
                "total_occurrences": max(markers, 1),
                "firm_occurrences": max(markers, 1),
                "conditional_occurrences": 0,
                "is_phone_call": False,
                "estimated_unit_cost": COST_MAP.get(tier, 75),
                "avg_confidence": 0.85,
                "source_pages": table.get("source_pages", []),
                "issues": [],
                "notes": "",
            })

    return lines


@app.get("/api/protocols/{protocol_id}/budget/export")
async def export_budget_xlsx(protocol_id: str):
    """Export site budget as a formatted XLSX file."""
    from fastapi.responses import Response

    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if not protocol:
        raise HTTPException(status_code=404, detail="Protocol not found")

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Site Budget"

        # Styles
        header_font = Font(bold=True, size=11, color="FFFFFF")
        header_fill = PatternFill(start_color="0093D0", fill_type="solid")
        cat_fill = PatternFill(start_color="F1F5F9", fill_type="solid")
        cat_font = Font(bold=True, size=10)
        total_fill = PatternFill(start_color="1E293B", fill_type="solid")
        total_font = Font(bold=True, size=11, color="FFFFFF")
        currency_fmt = '"$"#,##0.00'
        pct_fmt = "0%"
        thin_border = Border(
            left=Side(style="thin", color="E2E8F0"),
            right=Side(style="thin", color="E2E8F0"),
            top=Side(style="thin", color="E2E8F0"),
            bottom=Side(style="thin", color="E2E8F0"),
        )

        # Title
        meta = protocol.metadata
        ws.merge_cells("A1:I1")
        ws["A1"] = f"Site Budget — {meta.short_title or meta.title or protocol.document_name}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Protocol: {meta.protocol_number or protocol_id} | Phase: {meta.phase} | Sponsor: {meta.sponsor}"
        ws["A2"].font = Font(size=10, color="64748B")
        ws.append([])

        # Headers
        headers = [
            "Procedure", "Canonical Name", "CPT Code", "Category",
            "Cost Tier", "Visits", "Firm Occ.", "Cond. Occ.",
            "Unit Cost", "Line Total", "Confidence",
        ]
        ws.append(headers)
        header_row = ws.max_row
        for col_idx, _ in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", wrap_text=True)
            cell.border = thin_border

        # Budget lines grouped by category
        budget_lines = protocol.budget_lines or []
        categories: dict[str, list] = {}
        for bl in budget_lines:
            cat = bl.get("category", "Uncategorized") if isinstance(bl, dict) else getattr(bl, "category", "Uncategorized")
            if cat not in categories:
                categories[cat] = []
            categories[cat].append(bl)

        grand_total = 0
        for cat_name in sorted(categories.keys()):
            lines = categories[cat_name]
            # Category header
            ws.append([cat_name.upper()])
            cat_row = ws.max_row
            ws.merge_cells(f"A{cat_row}:K{cat_row}")
            ws.cell(row=cat_row, column=1).font = cat_font
            ws.cell(row=cat_row, column=1).fill = cat_fill

            cat_total = 0
            for bl in lines:
                if isinstance(bl, dict):
                    proc = bl.get("procedure", "")
                    canonical = bl.get("canonical_name", "")
                    cpt = bl.get("cpt_code", "")
                    category = bl.get("category", "")
                    tier = bl.get("cost_tier", "")
                    visits = len(bl.get("visits_required", []))
                    firm = bl.get("firm_occurrences", visits)
                    cond = bl.get("conditional_occurrences", 0)
                    unit_cost = bl.get("estimated_unit_cost", 0)
                    conf = bl.get("avg_confidence", 0)
                else:
                    proc = bl.procedure
                    canonical = bl.canonical_name
                    cpt = bl.cpt_code
                    category = bl.category
                    tier = bl.cost_tier
                    visits = len(bl.visits_required)
                    firm = getattr(bl, "firm_occurrences", visits)
                    cond = getattr(bl, "conditional_occurrences", 0)
                    unit_cost = bl.estimated_unit_cost
                    conf = bl.avg_confidence

                total_occ = firm + cond
                line_total = unit_cost * total_occ
                cat_total += line_total

                ws.append([
                    proc, canonical, cpt, category, tier,
                    total_occ, firm, cond, unit_cost, line_total, conf,
                ])
                row = ws.max_row
                ws.cell(row=row, column=9).number_format = currency_fmt
                ws.cell(row=row, column=10).number_format = currency_fmt
                ws.cell(row=row, column=11).number_format = pct_fmt
                for c in range(1, 12):
                    ws.cell(row=row, column=c).border = thin_border

            # Category subtotal
            ws.append(["", "", "", "", "", "", "", "Subtotal:", "", cat_total, ""])
            sub_row = ws.max_row
            ws.cell(row=sub_row, column=10).number_format = currency_fmt
            ws.cell(row=sub_row, column=10).font = Font(bold=True)
            grand_total += cat_total

        # Grand total
        ws.append([])
        ws.append(["", "", "", "", "", "", "", "GRAND TOTAL (Per Patient):", "", grand_total, ""])
        total_row = ws.max_row
        for c in range(1, 12):
            ws.cell(row=total_row, column=c).fill = total_fill
            ws.cell(row=total_row, column=c).font = total_font
        ws.cell(row=total_row, column=10).number_format = currency_fmt

        # Column widths
        widths = [30, 25, 10, 12, 10, 8, 8, 8, 12, 12, 10]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[chr(64 + i)].width = w

        # Save to bytes
        from io import BytesIO
        buf = BytesIO()
        wb.save(buf)
        buf.seek(0)

        fname = f"{protocol_id}_site_budget.xlsx"
        return Response(
            content=buf.getvalue(),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.document",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl not installed")
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Export failed: {e}")


@app.put("/api/procedures/{canonical_name}")
async def update_procedure(canonical_name: str, updates: dict):
    """Update a procedure's CPT code, category, cost tier, or aliases."""
    from src.domain.vocabulary import get_procedure_vocab

    vocab = get_procedure_vocab()
    entry = vocab.lookup(canonical_name)
    if not entry:
        raise HTTPException(
            status_code=404,
            detail=f"Procedure '{canonical_name}' not found",
        )

    if "cpt_code" in updates:
        entry.cpt_code = str(updates["cpt_code"])
    if "category" in updates:
        entry.category = str(updates["category"])
    if "cost_tier" in updates:
        entry.cost_tier = str(updates["cost_tier"])
    if "aliases" in updates:
        if isinstance(updates["aliases"], list):
            entry.aliases = updates["aliases"]
        elif isinstance(updates["aliases"], str):
            entry.aliases = [
                a.strip() for a in updates["aliases"].split(",") if a.strip()
            ]

    vocab._save()
    return {"status": "updated", "canonical_name": entry.canonical_name}


@app.delete("/api/procedures/{canonical_name}")
async def delete_procedure(canonical_name: str):
    """Delete a procedure from the library."""
    from src.domain.vocabulary import get_procedure_vocab

    vocab = get_procedure_vocab()
    key = canonical_name.lower()
    if key not in vocab._entries:
        raise HTTPException(
            status_code=404,
            detail=f"Procedure '{canonical_name}' not found",
        )

    del vocab._entries[key]
    # Remove from alias index
    to_remove = [
        alias for alias, canon in vocab._alias_index.items()
        if canon == key
    ]
    for alias in to_remove:
        del vocab._alias_index[alias]

    vocab._save()
    return {"status": "deleted", "canonical_name": canonical_name}


# ---------------------------------------------------------------------------
# Protocol & Knowledge Element endpoints
# ---------------------------------------------------------------------------


@app.get("/api/protocols")
async def list_protocols():
    """List all stored protocols."""
    store = create_ke_store()
    return store.list_protocols()


@app.get("/api/protocols/{protocol_id}")
async def get_protocol(protocol_id: str):
    """Get full protocol data."""
    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if not protocol:
        raise HTTPException(
            status_code=404, detail=f"Protocol {protocol_id} not found"
        )
    return protocol.model_dump(mode="json")


# ── Trust Endpoints ──────────────────────────────────────────────────────

_trust_cache: dict[str, dict] = {}  # protocol_id → ProtocolTrust dict


@app.get("/api/protocols/{protocol_id}/trust")
async def get_protocol_trust(protocol_id: str):
    """Compute and return protocol-level trust dashboard."""
    from src.trust.engine import compute_protocol_trust
    from src.models.schema import ExtractedTable

    # Check cache (invalidated on review actions)
    if protocol_id in _trust_cache:
        return _trust_cache[protocol_id]

    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if not protocol:
        raise HTTPException(status_code=404, detail=f"Protocol {protocol_id} not found")

    # Extract tables from protocol data
    protocol_data = protocol.model_dump(mode="json") if hasattr(protocol, "model_dump") else protocol
    tables_data = protocol_data.get("tables", [])

    # Reconstruct ExtractedTable objects for trust computation
    tables = []
    for td in tables_data:
        try:
            tables.append(ExtractedTable.model_validate(td))
        except Exception:
            continue

    trust = compute_protocol_trust(tables)
    result = trust.model_dump(mode="json")
    _trust_cache[protocol_id] = result
    return result


@app.get("/api/protocols/{protocol_id}/trust/rows")
async def get_protocol_trust_rows(protocol_id: str):
    """Get row-level trust breakdown for all procedures."""
    from src.trust.engine import compute_row_trust

    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if not protocol:
        raise HTTPException(status_code=404, detail=f"Protocol {protocol_id} not found")

    protocol_data = protocol.model_dump(mode="json") if hasattr(protocol, "model_dump") else protocol
    tables_data = protocol_data.get("tables", [])

    row_trusts = []
    for td in tables_data:
        procedures = {p.get("raw_name", ""): p for p in td.get("procedures", [])}
        cells = td.get("cells", [])
        footnotes = td.get("footnotes", [])

        # Group cells by row_header
        row_groups: dict[str, list[dict]] = {}
        for cell in cells:
            rh = cell.get("row_header", "")
            if rh:
                row_groups.setdefault(rh, []).append(cell)

        for proc_name, proc_cells in row_groups.items():
            proc_info = procedures.get(proc_name, {})
            confs = [c.get("confidence", 0.5) for c in proc_cells if c.get("col", 0) > 0]
            flagged = sum(1 for c in confs if c < 0.75)

            # Count footnote markers
            fn_total = sum(len(c.get("footnote_markers", [])) for c in proc_cells)
            fn_resolved = sum(len(c.get("resolved_footnotes", [])) for c in proc_cells)

            cpt = proc_info.get("code")
            category = proc_info.get("category", "")
            is_effort = category in ("Administrative", "Safety", "PRO")

            rt = compute_row_trust(
                procedure_name=proc_name,
                cell_confidences=confs if confs else [0.5],
                match_method="exact" if proc_info.get("canonical_name") else "unmatched",
                match_score=1.0 if proc_info.get("canonical_name") else 0.0,
                cpt_code=cpt,
                footnotes_total=fn_total,
                footnotes_resolved=fn_resolved,
                is_effort_based=is_effort,
                flagged_count=flagged,
            )
            row_trusts.append(rt.model_dump(mode="json"))

    return row_trusts


@app.get("/api/protocols/{protocol_id}/cells/{row}/{col}/evidence")
async def get_cell_evidence(protocol_id: str, row: int, col: int):
    """Get per-cell evidence chain for trust display."""
    from src.trust.models import CellEvidence

    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if not protocol:
        raise HTTPException(status_code=404, detail=f"Protocol {protocol_id} not found")

    protocol_data = protocol.model_dump(mode="json") if hasattr(protocol, "model_dump") else protocol

    # Search all tables for this cell
    for td in protocol_data.get("tables", []):
        for cell in td.get("cells", []):
            if cell.get("row") == row and cell.get("col") == col:
                evidence_data = cell.get("evidence")
                if evidence_data:
                    ev = CellEvidence.model_validate(evidence_data)
                    return {
                        "cell": {
                            "row": row, "col": col,
                            "raw_value": cell.get("raw_value", ""),
                            "data_type": cell.get("data_type", "TEXT"),
                            "confidence": cell.get("confidence", 0.5),
                            "row_header": cell.get("row_header", ""),
                            "col_header": cell.get("col_header", ""),
                            "footnote_markers": cell.get("footnote_markers", []),
                            "resolved_footnotes": cell.get("resolved_footnotes", []),
                        },
                        "evidence": ev.model_dump(mode="json"),
                        "verification_steps": [
                            s.model_dump(mode="json") for s in ev.verification_steps
                        ],
                        "challenger_issues": ev.challenger_issues,
                    }
                else:
                    # Legacy cell without evidence — return minimal response
                    return {
                        "cell": {
                            "row": row, "col": col,
                            "raw_value": cell.get("raw_value", ""),
                            "data_type": cell.get("data_type", "TEXT"),
                            "confidence": cell.get("confidence", 0.5),
                            "row_header": cell.get("row_header", ""),
                            "col_header": cell.get("col_header", ""),
                            "footnote_markers": cell.get("footnote_markers", []),
                            "resolved_footnotes": cell.get("resolved_footnotes", []),
                        },
                        "evidence": None,
                        "verification_steps": [],
                        "challenger_issues": [],
                    }

    raise HTTPException(status_code=404, detail=f"Cell ({row},{col}) not found")


@app.get("/api/protocols/{protocol_id}/sections")
async def get_protocol_sections(protocol_id: str):
    """Parse sections from a protocol's PDF. Used when stored sections are empty."""
    pdf_path = _find_protocol_pdf(protocol_id)
    if not pdf_path:
        raise HTTPException(status_code=404, detail="PDF not available for this protocol")

    from src.pipeline.section_parser import SectionParser
    parser = SectionParser()
    pdf_bytes = pdf_path.read_bytes()
    sections = parser.parse(pdf_bytes, filename=pdf_path.name)

    def serialize(secs):
        result = []
        for s in secs:
            result.append({
                "number": s.number,
                "title": s.title,
                "page": s.page,
                "end_page": s.end_page,
                "level": s.level,
                "children": serialize(s.children) if s.children else [],
            })
        return result

    return {
        "sections": serialize(sections),
        "method": "parsed_from_pdf",
        "total_sections": len(parser._flatten(sections)),
    }


@app.get("/api/protocols/{protocol_id}/page-image/{page_number}")
async def get_page_image(protocol_id: str, page_number: int):
    """Render a PDF page as a PNG image for the document viewer."""
    from fastapi.responses import Response

    pdf_path = _find_protocol_pdf(protocol_id)
    if not pdf_path:
        raise HTTPException(
            status_code=404,
            detail=f"PDF not found for {protocol_id}",
        )

    try:
        import fitz

        doc = fitz.open(str(pdf_path))
        if page_number < 0 or page_number >= doc.page_count:
            doc.close()
            raise HTTPException(
                status_code=404,
                detail=f"Page {page_number} out of range (0-{doc.page_count - 1})",
            )
        page = doc[page_number]
        pix = page.get_pixmap(dpi=150)
        img_bytes = pix.tobytes("png")
        doc.close()
        return Response(
            content=img_bytes,
            media_type="image/png",
            headers={
                "Cache-Control": "public, max-age=3600",
                "Access-Control-Allow-Origin": "*",
            },
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(
            status_code=500, detail=f"Failed to render page: {e}"
        )


@app.get("/api/protocols/{protocol_id}/sections")
async def get_protocol_sections(protocol_id: str):
    """Get the section tree for a protocol."""
    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if not protocol:
        raise HTTPException(
            status_code=404, detail=f"Protocol {protocol_id} not found"
        )
    return [s.model_dump() for s in protocol.sections]


@app.get("/api/protocols/{protocol_id}/sections/{section_number:path}")
async def get_section_content(protocol_id: str, section_number: str):
    """Get content for a specific section."""
    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if not protocol:
        raise HTTPException(
            status_code=404, detail=f"Protocol {protocol_id} not found"
        )

    def find_section(sections, number):
        for s in sections:
            if s.number == number:
                return s
            found = find_section(s.children, number)
            if found:
                return found
        return None

    section = find_section(protocol.sections, section_number)
    if not section:
        raise HTTPException(
            status_code=404, detail=f"Section {section_number} not found"
        )
    return section.model_dump()


@app.get("/api/protocols/{protocol_id}/budget")
async def get_protocol_budget(protocol_id: str):
    """Get budget line items for a protocol."""
    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if not protocol:
        raise HTTPException(
            status_code=404, detail=f"Protocol {protocol_id} not found"
        )
    return protocol.budget_lines


@app.get("/api/protocols/{protocol_id}/knowledge-elements")
async def get_knowledge_elements(
    protocol_id: str, ke_type: str | None = None
):
    """Get knowledge elements, optionally filtered by type."""
    store = create_ke_store()
    kes = store.get_knowledge_elements(protocol_id, ke_type)
    return [ke.model_dump() for ke in kes]


# ---------------------------------------------------------------------------
# SMB (Structured Model Builder) endpoints
# ---------------------------------------------------------------------------

_smb_cache: dict[str, dict[str, Any]] = {}  # protocol_id → BuildResult dict
_smb_building: set[str] = set()  # protocol IDs currently building


@app.post("/api/smb/build/{protocol_id}")
async def smb_build(protocol_id: str):
    """Build a structured model from a stored protocol."""
    from src.smb.core.engine import SMBEngine

    if protocol_id in _smb_building:
        return {"status": "building", "protocol_id": protocol_id}

    store = create_ke_store()
    protocol = store.load_protocol(protocol_id)
    if not protocol:
        raise HTTPException(
            status_code=404, detail=f"Protocol {protocol_id} not found"
        )

    _smb_building.add(protocol_id)
    try:
        protocol_data = (
            protocol.model_dump(mode="json")
            if hasattr(protocol, "model_dump")
            else protocol
        )
        engine = SMBEngine(domain="protocol")
        result = engine.build_from_protocol_json(protocol_data)

        # Cache the full result
        model_dict = result.model.model_dump(mode="json")
        graph_dict = result.model.to_graph_dict()
        summary = result.model.summary()

        # Import query functions for budget schedule
        from src.smb.core.query import (
            get_budget_schedule,
            get_visit_timeline,
        )

        schedule = get_budget_schedule(result.model)
        timeline = get_visit_timeline(result.model)

        _smb_cache[protocol_id] = {
            "model": model_dict,
            "graph": graph_dict,
            "summary": summary,
            "schedule": schedule,
            "timeline": timeline,
            "build_time_seconds": result.build_time_seconds,
            "inference_rules_fired": result.inference_rules_fired,
            "validation_passed": result.validation_passed,
            "validation_errors": result.validation_errors,
            "validation_warnings": result.validation_warnings,
        }

        return {
            "status": "ready",
            "protocol_id": protocol_id,
            "summary": summary,
            "build_time_seconds": result.build_time_seconds,
            "inference_rules_fired": result.inference_rules_fired,
            "validation_passed": result.validation_passed,
        }
    except Exception as e:
        logger.exception(f"SMB build failed for {protocol_id}: {e}")
        raise HTTPException(
            status_code=500,
            detail=f"SMB build failed: {str(e)}",
        )
    finally:
        _smb_building.discard(protocol_id)


@app.get("/api/smb/model/{protocol_id}")
async def smb_model_summary(protocol_id: str):
    """Get structured model summary for a protocol."""
    if protocol_id not in _smb_cache:
        raise HTTPException(
            status_code=404,
            detail=f"SMB model not built for {protocol_id}. POST /api/smb/build/{protocol_id} first.",
        )
    cached = _smb_cache[protocol_id]
    return {
        "protocol_id": protocol_id,
        "summary": cached["summary"],
        "build_time_seconds": cached["build_time_seconds"],
        "inference_rules_fired": cached["inference_rules_fired"],
        "validation_passed": cached["validation_passed"],
        "validation_errors": cached["validation_errors"],
        "validation_warnings": cached["validation_warnings"],
        "timeline": cached["timeline"],
    }


@app.get("/api/smb/model/{protocol_id}/schedule")
async def smb_model_schedule(protocol_id: str):
    """Get ScheduleEntry matrix formatted for the budget calculator."""
    if protocol_id not in _smb_cache:
        raise HTTPException(
            status_code=404,
            detail=f"SMB model not built for {protocol_id}. POST /api/smb/build/{protocol_id} first.",
        )
    return {
        "protocol_id": protocol_id,
        "schedule": _smb_cache[protocol_id]["schedule"],
    }


@app.get("/api/smb/model/{protocol_id}/graph")
async def smb_model_graph(protocol_id: str):
    """Get entity-relationship graph for visualization."""
    if protocol_id not in _smb_cache:
        raise HTTPException(
            status_code=404,
            detail=f"SMB model not built for {protocol_id}. POST /api/smb/build/{protocol_id} first.",
        )
    return _smb_cache[protocol_id]["graph"]


# ---------------------------------------------------------------------------
# Feedback System — submit, triage, track, delivery reports
# ---------------------------------------------------------------------------

FEEDBACK_DIR = Path("data/feedback")
FEEDBACK_FILE = FEEDBACK_DIR / "backlog.jsonl"


class FeedbackSubmission(BaseModel):
    """User-submitted feedback from the UI."""
    category: str = Field(..., pattern="^(bug|issue|enhancement|feature)$")
    title: str = Field(..., min_length=3, max_length=500)
    description: str = Field(..., min_length=3)
    priority: str = Field(default="medium", pattern="^(low|medium|high|critical)$")
    page_url: str = ""
    attachments: list[dict[str, str]] = Field(default_factory=list)


class FeedbackStatusUpdate(BaseModel):
    status: str = Field(..., pattern="^(new|triaging|spec_ready|in_progress|testing|deploying|delivered|rejected)$")
    resolution: str | None = None


def _load_feedback() -> list[dict[str, Any]]:
    """Load all feedback entries from JSONL."""
    if not FEEDBACK_FILE.exists():
        return []
    entries = []
    for line in FEEDBACK_FILE.read_text(encoding="utf-8").strip().split("\n"):
        if line.strip():
            try:
                entries.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return entries


def _save_feedback_entry(entry: dict[str, Any]) -> None:
    """Append a single feedback entry to the JSONL file."""
    FEEDBACK_DIR.mkdir(parents=True, exist_ok=True)
    with open(FEEDBACK_FILE, "a", encoding="utf-8") as f:
        f.write(json.dumps(entry, default=str) + "\n")


def _update_feedback_entry(entry_id: str, updates: dict[str, Any]) -> dict[str, Any] | None:
    """Update a feedback entry in place. Rewrites the JSONL."""
    entries = _load_feedback()
    updated = None
    for e in entries:
        if e["id"] == entry_id:
            e.update(updates)
            e["updated_at"] = time.time()
            updated = e
            break
    if updated:
        FEEDBACK_DIR.mkdir(parents=True, exist_ok=True)
        with open(FEEDBACK_FILE, "w", encoding="utf-8") as f:
            for e in entries:
                f.write(json.dumps(e, default=str) + "\n")
    return updated


@app.post("/api/feedback")
async def submit_feedback(submission: FeedbackSubmission):
    """Submit user feedback. Stored for pickup by Claude Code session which
    has full repo context for triage, spec, TDD, implementation, and deploy."""
    entry_id = str(uuid.uuid4())[:8]
    now = time.time()

    entry = {
        "id": entry_id,
        "submitted_at": now,
        "updated_at": now,
        "category": submission.category,
        "title": submission.title,
        "description": submission.description,
        "priority": submission.priority,
        "page_url": submission.page_url,
        "attachments": submission.attachments[:5],  # max 5
        "status": "new",
        "triage": None,
        "delivery_report": None,
    }

    _save_feedback_entry(entry)
    logger.info(f"Feedback submitted: {entry_id} — {submission.title}")

    return {
        "id": entry_id,
        "status": "new",
        "message": "Feedback submitted. Will be picked up by the development agent.",
    }


@app.get("/api/feedback")
async def list_feedback(status: str | None = None, limit: int = 50, offset: int = 0):
    """List feedback entries, newest first. Optionally filter by status."""
    entries = _load_feedback()
    if status:
        entries = [e for e in entries if e.get("status") == status]
    entries.sort(key=lambda e: e.get("submitted_at", 0), reverse=True)
    total = len(entries)
    entries = entries[offset:offset + limit]
    return {"items": entries, "total": total}


@app.get("/api/feedback/{entry_id}")
async def get_feedback(entry_id: str):
    """Get a single feedback entry with full triage, spec, and delivery report."""
    entries = _load_feedback()
    for e in entries:
        if e["id"] == entry_id:
            return e
    raise HTTPException(status_code=404, detail=f"Feedback {entry_id} not found")


@app.patch("/api/feedback/{entry_id}")
async def update_feedback_status(entry_id: str, update: FeedbackStatusUpdate):
    """Update feedback status and optional resolution/delivery report."""
    updates: dict[str, Any] = {"status": update.status}
    if update.resolution:
        updates["resolution"] = update.resolution
    if update.status == "delivered":
        updates["delivered_at"] = time.time()
        # Auto-generate delivery report
        entry = None
        for e in _load_feedback():
            if e["id"] == entry_id:
                entry = e
                break
        if entry and entry.get("triage"):
            updates["delivery_report"] = {
                "title": entry["title"],
                "status": "delivered",
                "triage_summary": entry["triage"].get("root_cause_hypothesis", ""),
                "fix_applied": update.resolution or entry["triage"].get("suggested_fix", ""),
                "spec": entry["triage"].get("spec", {}),
                "tdd_plan": entry["triage"].get("tdd_plan", {}),
                "delivered_at": time.time(),
            }

    result = _update_feedback_entry(entry_id, updates)
    if not result:
        raise HTTPException(status_code=404, detail=f"Feedback {entry_id} not found")
    return result


# ---------------------------------------------------------------------------
# Document Fidelity Checker endpoints
# ---------------------------------------------------------------------------

@app.post("/api/fidelity/check")
async def fidelity_check(file: UploadFile = File(...)):
    """Check a PDF document for formatting fidelity issues.

    Detects run-on words, spacing problems, font inconsistencies,
    alignment issues, and color anomalies.
    """
    from src.formatter.fidelity_checker import DocumentFidelityChecker

    pdf_bytes = await file.read()
    checker = DocumentFidelityChecker()
    report = checker.check(pdf_bytes, filename=file.filename or "")

    return {
        "document_name": report.document_name,
        "score": round(report.score, 1),
        "total_issues": report.total_issues,
        "critical": report.critical_count,
        "high": report.high_count,
        "medium": report.medium_count,
        "low": report.low_count,
        "issues": [
            {
                "category": i.category,
                "severity": i.severity,
                "page": i.page,
                "location": i.location,
                "description": i.description,
                "original_text": i.original_text,
                "suggested_fix": i.suggested_fix,
                "auto_fixable": i.auto_fixable,
            }
            for i in report.issues
        ],
        "formatting_summary": report.formatting_summary,
    }


@app.post("/api/fidelity/compare")
async def fidelity_compare(
    template: UploadFile = File(...),
    generated: UploadFile = File(...),
):
    """Compare a generated document against a template for formatting conformance.

    Upload two PDFs: the blueprint/template and the generated document.
    Returns a fidelity report showing where the generated doc deviates.
    """
    from src.formatter.fidelity_checker import DocumentFidelityChecker

    template_bytes = await template.read()
    generated_bytes = await generated.read()

    checker = DocumentFidelityChecker()
    report = checker.compare(
        template_bytes, generated_bytes,
        template_name=template.filename or "template",
        generated_name=generated.filename or "generated",
    )

    return {
        "template_name": template.filename,
        "generated_name": generated.filename,
        "score": round(report.score, 1),
        "total_issues": report.total_issues,
        "critical": report.critical_count,
        "high": report.high_count,
        "medium": report.medium_count,
        "low": report.low_count,
        "issues": [
            {
                "category": i.category,
                "severity": i.severity,
                "page": i.page,
                "location": i.location,
                "description": i.description,
                "original_text": i.original_text,
                "suggested_fix": i.suggested_fix,
                "auto_fixable": i.auto_fixable,
            }
            for i in report.issues
        ],
        "formatting_summary": report.formatting_summary,
    }


@app.post("/api/fidelity/extract-formatted")
async def extract_formatted(file: UploadFile = File(...)):
    """Extract a PDF with full formatting preserved.

    Returns the document as styled HTML (suitable for CKEditor or preview)
    along with formatting metadata (fonts, colors, styles used).
    """
    from src.formatter.extractor import FormattingExtractor

    pdf_bytes = await file.read()
    extractor = FormattingExtractor()
    doc = extractor.extract(pdf_bytes, filename=file.filename or "")

    # Render to HTML with inline styles
    html_parts = []
    for page in doc.pages:
        html_parts.append(f'<div class="page" data-page="{page.page_number + 1}">')
        for para in page.paragraphs:
            tag = "p"
            if para.style.startswith("heading"):
                level = para.style[-1] if para.style[-1].isdigit() else "3"
                tag = f"h{level}"
            elif para.style == "list_bullet":
                tag = "li"
            elif para.style == "list_number":
                tag = "li"

            # Build inline HTML from spans
            spans_html = []
            for line in para.lines:
                for span in line.spans:
                    text = span.text.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;")
                    styles = []
                    if span.font:
                        styles.append(f"font-family:'{span.font_family}'")
                    if span.size:
                        styles.append(f"font-size:{span.size:.1f}pt")
                    r, g, b = span.color_rgb
                    if r > 30 or g > 30 or b > 30:
                        styles.append(f"color:rgb({r},{g},{b})")
                    if span.bold:
                        text = f"<strong>{text}</strong>"
                    if span.italic:
                        text = f"<em>{text}</em>"
                    if span.underline:
                        text = f"<u>{text}</u>"
                    if span.superscript:
                        text = f"<sup>{text}</sup>"
                    if span.subscript:
                        text = f"<sub>{text}</sub>"

                    if styles:
                        text = f'<span style="{";".join(styles)}">{text}</span>'
                    spans_html.append(text)

            content = "".join(spans_html)
            indent_style = f"margin-left:{para.indent_level * 24}px;" if para.indent_level else ""
            align_style = f"text-align:{para.alignment};" if para.alignment != "left" else ""
            style_attr = f' style="{indent_style}{align_style}"' if indent_style or align_style else ""

            html_parts.append(f"<{tag}{style_attr}>{content}</{tag}>")
        html_parts.append("</div>")

    return {
        "document_name": file.filename,
        "html": "\n".join(html_parts),
        "total_pages": len(doc.pages),
        "total_paragraphs": doc.total_paragraphs,
        "fonts_used": dict(sorted(doc.font_inventory.items(), key=lambda x: -x[1])[:10]),
        "colors_used": dict(sorted(doc.color_inventory.items(), key=lambda x: -x[1])[:10]),
        "styles_used": doc.style_inventory,
    }


# ---------------------------------------------------------------------------
# Template-Guided Document Generator endpoint
# ---------------------------------------------------------------------------

@app.post("/api/fidelity/generate")
async def template_generate(
    template: UploadFile = File(...),
    source: UploadFile = File(...),
):
    """Generate a document by applying template formatting to source content.

    Accepts two PDF files:
    - template: the blueprint PDF whose formatting will be extracted
    - source: the input PDF whose content will be re-styled

    Returns generated HTML, a style profile, and a conformance report.
    """
    from src.formatter.template_generator import TemplateConformer

    template_bytes = await template.read()
    source_bytes = await source.read()

    conformer = TemplateConformer()
    result = conformer.conform(template_bytes, source_bytes)

    return {
        "template_name": template.filename,
        "source_name": source.filename,
        "html": result["html"],
        "style_profile": result["style_profile"],
        "conformance_report": result["conformance_report"],
    }


@app.post("/api/fidelity/export-docx")
async def export_docx(file: UploadFile = File(...)):
    """Convert a PDF to a formatting-preserved DOCX document.

    Extracts all formatting (fonts, sizes, colors, bold/italic,
    super/subscript) from the PDF and renders to a Word document.
    """
    from src.formatter.docx_renderer import DOCXRenderer
    from fastapi.responses import Response

    pdf_bytes = await file.read()
    renderer = DOCXRenderer()
    docx_bytes = renderer.render_from_pdf(pdf_bytes, filename=file.filename or "")

    output_name = (file.filename or "document").replace(".pdf", ".docx")
    return Response(
        content=docx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f'attachment; filename="{output_name}"',
            "Access-Control-Allow-Origin": "*",
        },
    )


@app.post("/api/fidelity/export-docx-from-template")
async def export_docx_from_template(
    template: UploadFile = File(...),
    source: UploadFile = File(...),
):
    """Generate a DOCX by applying template formatting to source content.

    Takes a blueprint template PDF (formatting source) and a source PDF
    (content source), produces a DOCX with template formatting applied.
    """
    from src.formatter.docx_renderer import DOCXRenderer
    from src.formatter.template_generator import TemplateConformer
    from fastapi.responses import Response

    template_bytes = await template.read()
    source_bytes = await source.read()

    conformer = TemplateConformer()
    profile = conformer.extract_style_profile(template_bytes)
    source_doc = conformer.extractor.extract(source_bytes)

    renderer = DOCXRenderer()
    docx_bytes = renderer.render_with_profile(source_doc, profile.to_dict())

    output_name = (source.filename or "document").replace(".pdf", "_formatted.docx")
    return Response(
        content=docx_bytes,
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": f'attachment; filename="{output_name}"',
            "Access-Control-Allow-Origin": "*",
        },
    )


@app.post("/api/fidelity/detect-formulas")
async def detect_formulas(file: UploadFile = File(...)):
    """Detect formulas in a document using the registry-based formula system.

    Supports PDF, DOCX, HTML, PPTX, XLSX input.
    Returns formulas with type, original text, HTML, LaTeX, complexity tier, and source.
    """
    from src.formatter.formula.factory import create_formula_system
    from src.formatter import DocHandler

    content = await file.read()
    filename = file.filename or ""

    # Detect format from extension
    ext = filename.rsplit(".", 1)[-1].lower() if "." in filename else "pdf"
    format_map = {"pdf": "pdf", "docx": "docx", "html": "html", "htm": "html",
                  "pptx": "pptx", "xlsx": "xlsx", "md": "markdown", "txt": "text"}
    fmt = format_map.get(ext, "pdf")

    # Ingest document
    handler = DocHandler()
    doc = handler.ingest(content, fmt, filename)

    # Run formula detection
    orchestrator = create_formula_system()
    all_formulas = []
    formula_counts: dict[str, int] = {}
    tier_counts = {"inline": 0, "structured": 0, "rendered": 0}

    for page in doc.pages:
        for para_idx, para in enumerate(page.paragraphs):
            text = para.text
            if not text.strip():
                continue
            spans = orchestrator.process_text(text)
            for s in spans:
                f = s.formula
                all_formulas.append({
                    "page": page.page_number + 1,
                    "paragraph": para_idx + 1,
                    "type": f.formula_type.value,
                    "original": s.original_text,
                    "html": f.html or s.original_text,
                    "latex": f.latex or "",
                    "complexity": f.complexity.value,
                    "source": f.source.value,
                    "confidence": f.confidence,
                })
                ftype = f.formula_type.value
                formula_counts[ftype] = formula_counts.get(ftype, 0) + 1
                tier_counts[f.complexity.value] = tier_counts.get(f.complexity.value, 0) + 1

    return {
        "document_name": filename,
        "total_formulas": len(all_formulas),
        "by_type": formula_counts,
        "by_tier": tier_counts,
        "registry_tools": orchestrator._registry.list_tools(),
        "formulas": all_formulas,
    }


@app.post("/api/fidelity/generate-contract")
async def generate_site_contract(
    template: UploadFile = File(...),
    protocol_id: str | None = None,
):
    """Generate a site contract from a CTSA template + stored protocol data.

    Upload a CTSA template PDF. If protocol_id is provided, fills with
    that protocol's extracted data. Returns HTML preview + DOCX download.
    """
    from src.formatter.site_contract_generator import SiteContractGenerator

    template_bytes = await template.read()

    # Load protocol data
    protocol_data: dict[str, Any] = {}
    if protocol_id:
        store = create_ke_store()
        protocol = store.load_protocol(protocol_id)
        if protocol:
            protocol_data = (
                protocol.model_dump(mode="json")
                if hasattr(protocol, "model_dump")
                else protocol
            )

    generator = SiteContractGenerator()
    result = generator.generate(template_bytes, protocol_data)

    return {
        "html": result["html"],
        "fill_report": result["fill_report"],
        "template_pages": result["template_pages"],
        "template_paragraphs": result["template_paragraphs"],
    }


@app.post("/api/fidelity/generate-contract-docx")
async def generate_site_contract_docx(
    template: UploadFile = File(...),
    protocol_id: str | None = None,
):
    """Generate a site contract DOCX from a CTSA template + protocol data."""
    from src.formatter.site_contract_generator import SiteContractGenerator
    from fastapi.responses import Response

    template_bytes = await template.read()

    protocol_data: dict[str, Any] = {}
    if protocol_id:
        store = create_ke_store()
        protocol = store.load_protocol(protocol_id)
        if protocol:
            protocol_data = (
                protocol.model_dump(mode="json")
                if hasattr(protocol, "model_dump")
                else protocol
            )

    generator = SiteContractGenerator()
    result = generator.generate(template_bytes, protocol_data)

    return Response(
        content=result["docx_bytes"],
        media_type="application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        headers={
            "Content-Disposition": 'attachment; filename="site_contract.docx"',
            "Access-Control-Allow-Origin": "*",
        },
    )


if __name__ == "__main__":
    import uvicorn
    port = int(os.environ.get("PORT", 8000))
    uvicorn.run(app, host="0.0.0.0", port=port)
