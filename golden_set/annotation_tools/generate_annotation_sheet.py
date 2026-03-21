"""
Annotation Sheet Generator — creates a structured Excel workbook for
ground truth annotation from pipeline extraction output.

Generates 4 tabs:
1. soa_cells: Cell-level verification (is_correct + correct_value)
2. footnotes: Footnote binding verification
3. procedures: Procedure mapping verification
4. sections: Section parsing verification

The annotator opens this alongside the source PDF. For the ~90% of cells
that are correct, a single keystroke ("Y") in is_correct is sufficient.
Only the ~10% that need correction require typing.

Output is designed to convert cleanly to OmniDocBench JSON format.

Usage:
    python generate_annotation_sheet.py extraction_result.json output.xlsx
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side, numbers
)
from openpyxl.utils import get_column_letter


# Colors
HEADER_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
HEADER_FONT = Font(name="Segoe UI", size=10, bold=True, color="FFFFFF")
CELL_FONT = Font(name="Segoe UI", size=9)
CORRECT_FILL = PatternFill(start_color="ECFDF5", end_color="ECFDF5", fill_type="solid")
WRONG_FILL = PatternFill(start_color="FEF2F2", end_color="FEF2F2", fill_type="solid")
INPUT_FILL = PatternFill(start_color="FEFCE8", end_color="FEFCE8", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="E2E8F0"),
    right=Side(style="thin", color="E2E8F0"),
    top=Side(style="thin", color="E2E8F0"),
    bottom=Side(style="thin", color="E2E8F0"),
)


def generate_annotation_workbook(result: dict, output_path: Path):
    """Generate the annotation Excel workbook from pipeline output."""
    wb = Workbook()

    # Tab 1: SoA Cells
    _build_cells_tab(wb, result)

    # Tab 2: Footnotes
    _build_footnotes_tab(wb, result)

    # Tab 3: Procedures
    _build_procedures_tab(wb, result)

    # Tab 4: Sections (if section data available)
    _build_sections_tab(wb, result)

    # Tab 5: Instructions
    _build_instructions_tab(wb)

    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    wb.save(output_path)
    print(f"Annotation workbook saved: {output_path}")


def _build_cells_tab(wb: Workbook, result: dict):
    ws = wb.create_sheet("soa_cells", 0)

    headers = [
        "table_id", "source_pages", "row", "col", "row_header", "col_header",
        "extracted_value", "data_type", "confidence",
        "is_correct (Y/N)", "correct_value", "footnote_markers",
        "difficulty", "notes"
    ]
    _write_header_row(ws, headers)

    # Annotator input columns (highlighted yellow)
    input_cols = {10, 11, 13, 14}  # is_correct, correct_value, difficulty, notes

    row_num = 2
    for table in result.get("tables", []):
        table_id = table.get("table_id", "")
        pages = ", ".join(str(p) for p in table.get("source_pages", []))
        flagged = {(f["row"], f["col"]) for f in table.get("flagged_cells", [])}

        for cell in table.get("cells", []):
            is_flagged = (cell["row"], cell["col"]) in flagged
            ws.cell(row=row_num, column=1, value=table_id)
            ws.cell(row=row_num, column=2, value=pages)
            ws.cell(row=row_num, column=3, value=cell["row"])
            ws.cell(row=row_num, column=4, value=cell["col"])
            ws.cell(row=row_num, column=5, value=cell.get("row_header", ""))
            ws.cell(row=row_num, column=6, value=cell.get("col_header", ""))
            ws.cell(row=row_num, column=7, value=cell.get("raw_value", ""))
            ws.cell(row=row_num, column=8, value=cell.get("data_type", ""))
            ws.cell(row=row_num, column=9, value=round(cell.get("confidence", 0), 2))

            # DO NOT pre-fill is_correct — confidence is the model's
            # self-assessment, not ground truth. Every cell must be verified
            # by a human against the source PDF. Pre-filling Y creates
            # false ground truth that undermines the entire evaluation.
            ws.cell(row=row_num, column=10, value="")

            ws.cell(row=row_num, column=11, value="")  # correct_value
            ws.cell(row=row_num, column=12, value=", ".join(cell.get("footnote_markers", [])))
            ws.cell(row=row_num, column=13, value="")  # difficulty
            ws.cell(row=row_num, column=14, value="")  # notes

            # Style
            for col in range(1, 15):
                c = ws.cell(row=row_num, column=col)
                c.font = CELL_FONT
                c.border = THIN_BORDER
                c.alignment = Alignment(vertical="top", wrap_text=True)
                if col in input_cols:
                    c.fill = INPUT_FILL
                elif is_flagged:
                    c.fill = WRONG_FILL

            row_num += 1

    # Column widths
    widths = [12, 10, 5, 5, 30, 20, 20, 10, 8, 12, 20, 12, 10, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _build_footnotes_tab(wb: Workbook, result: dict):
    ws = wb.create_sheet("footnotes")

    headers = [
        "table_id", "marker", "extracted_text", "extracted_type",
        "bound_cells (row:col pairs)", "cell_count",
        "is_correct (Y/N)", "correct_text", "correct_type",
        "correct_bindings", "notes"
    ]
    _write_header_row(ws, headers)

    input_cols = {7, 8, 9, 10, 11}

    row_num = 2
    for table in result.get("tables", []):
        for fn in table.get("footnotes", []):
            bindings = ", ".join(
                f"{c['row']}:{c['col']}" for c in fn.get("applies_to", [])
            )
            ws.cell(row=row_num, column=1, value=table.get("table_id", ""))
            ws.cell(row=row_num, column=2, value=fn.get("marker", ""))
            ws.cell(row=row_num, column=3, value=fn.get("text", ""))
            ws.cell(row=row_num, column=4, value=fn.get("footnote_type", ""))
            ws.cell(row=row_num, column=5, value=bindings)
            ws.cell(row=row_num, column=6, value=len(fn.get("applies_to", [])))
            ws.cell(row=row_num, column=7, value="")  # is_correct
            ws.cell(row=row_num, column=8, value="")  # correct_text
            ws.cell(row=row_num, column=9, value="")  # correct_type
            ws.cell(row=row_num, column=10, value="")  # correct_bindings
            ws.cell(row=row_num, column=11, value="")  # notes

            for col in range(1, 12):
                c = ws.cell(row=row_num, column=col)
                c.font = CELL_FONT
                c.border = THIN_BORDER
                c.alignment = Alignment(vertical="top", wrap_text=True)
                if col in input_cols:
                    c.fill = INPUT_FILL

            row_num += 1

    widths = [12, 6, 50, 15, 30, 8, 12, 50, 15, 30, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def _build_procedures_tab(wb: Workbook, result: dict):
    ws = wb.create_sheet("procedures")

    headers = [
        "extracted_name", "canonical_name", "cpt_code", "code_system",
        "category", "cost_tier",
        "is_correct (Y/N)", "correct_canonical", "correct_cpt", "notes"
    ]
    _write_header_row(ws, headers)

    input_cols = {7, 8, 9, 10}
    seen = set()

    row_num = 2
    for table in result.get("tables", []):
        for proc in table.get("procedures", []):
            key = proc.get("raw_name", "")
            if key in seen:
                continue
            seen.add(key)

            ws.cell(row=row_num, column=1, value=proc.get("raw_name", ""))
            ws.cell(row=row_num, column=2, value=proc.get("canonical_name", ""))
            ws.cell(row=row_num, column=3, value=proc.get("code", ""))
            ws.cell(row=row_num, column=4, value=proc.get("code_system", ""))
            ws.cell(row=row_num, column=5, value=proc.get("category", ""))
            ws.cell(row=row_num, column=6, value=proc.get("estimated_cost_tier", ""))
            ws.cell(row=row_num, column=7, value="")
            ws.cell(row=row_num, column=8, value="")
            ws.cell(row=row_num, column=9, value="")
            ws.cell(row=row_num, column=10, value="")

            for col in range(1, 11):
                c = ws.cell(row=row_num, column=col)
                c.font = CELL_FONT
                c.border = THIN_BORDER
                c.alignment = Alignment(vertical="top", wrap_text=True)
                if col in input_cols:
                    c.fill = INPUT_FILL

            row_num += 1

    widths = [40, 35, 10, 8, 15, 10, 12, 35, 10, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def _build_sections_tab(wb: Workbook, result: dict):
    """Build sections tab — requires section data from the parser."""
    ws = wb.create_sheet("sections")

    headers = [
        "section_number", "extracted_title", "extracted_page",
        "is_correct (Y/N)", "correct_title", "correct_page", "notes"
    ]
    _write_header_row(ws, headers)

    # Placeholder — sections come from the section parser, not extraction output
    ws.cell(row=2, column=1, value="(Run section parser separately and paste results here)")
    ws.cell(row=2, column=1).font = Font(italic=True, color="94A3B8")

    widths = [15, 50, 12, 12, 50, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"


def _build_instructions_tab(wb: Workbook):
    ws = wb.create_sheet("instructions")

    instructions = [
        ("HOW TO ANNOTATE", "", True),
        ("", "", False),
        ("1. Open the source protocol PDF alongside this spreadsheet.", "", False),
        ("2. For each row in the soa_cells tab:", "", False),
        ("   - Compare 'extracted_value' against what you see in the PDF", "", False),
        ("   - If correct: type Y in 'is_correct'", "", False),
        ("   - If wrong: type N in 'is_correct' and enter the correct value in 'correct_value'", "", False),
        ("   - Optionally set difficulty: trivial / moderate / hard", "", False),
        ("3. Yellow columns are YOUR input. White columns are pipeline output.", "", False),
        ("4. Red-highlighted rows were flagged by the pipeline as uncertain.", "", False),
        ("", "", False),
        ("FOOTNOTES TAB", "", True),
        ("- Verify each footnote's text matches the source", "", False),
        ("- Verify the bound_cells list — does footnote 'a' apply to those cells?", "", False),
        ("- Verify the type classification (CONDITIONAL / EXCEPTION / REFERENCE / CLARIFICATION)", "", False),
        ("", "", False),
        ("PROCEDURES TAB", "", True),
        ("- Verify canonical_name is the correct standardized name", "", False),
        ("- Verify CPT code is correct for this procedure", "", False),
        ("- Verify cost_tier reflects the actual cost level", "", False),
        ("", "", False),
        ("AFTER ANNOTATION", "", True),
        ("Run: python golden_set/annotation_tools/convert_to_json.py annotation.xlsx output.json", "", False),
        ("This converts your annotations to OmniDocBench-compatible JSON for TEDS computation.", "", False),
    ]

    for i, (text, _, bold) in enumerate(instructions, 1):
        c = ws.cell(row=i, column=1, value=text)
        if bold:
            c.font = Font(name="Segoe UI", size=11, bold=True)
        else:
            c.font = Font(name="Segoe UI", size=10)

    ws.column_dimensions["A"].width = 80


def _write_header_row(ws, headers: list[str]):
    for i, header in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=header)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = THIN_BORDER


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python generate_annotation_sheet.py <extraction.json> <output.xlsx>")
        sys.exit(1)

    with open(sys.argv[1], encoding="utf-8") as f:
        result = json.load(f)

    generate_annotation_workbook(result, Path(sys.argv[2]))
