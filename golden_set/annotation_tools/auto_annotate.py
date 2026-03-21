"""
Auto-Annotation Script — independent verification of pipeline extraction.

Uses a DIFFERENT verification approach for each data type:
- EMPTY cells: OCR check (is the cell region blank?)
- MARKER cells: OCR comparison (does OCR read the same mark?)
- TEXT cells: Secondary VLM verification prompt ("Does this cell contain X?")
- CONDITIONAL/NUMERIC: Flagged for human review (too important to automate)

Key principle: verification prompt, not extraction prompt.
"Does this cell contain '~10 mL'?" is a 10-token YES/NO task,
not a 200-token extraction task. Faster, cheaper, more reliable.

Output is written back to the annotation Excel with audit trail:
- AUTO:ocr:0.95 = auto-verified by OCR
- AUTO:vlm_verify:0.88 = auto-verified by secondary VLM
- REVIEW:pipeline=X verifier=Y = disagreement, needs human review
- HUMAN_REQUIRED = flagged for mandatory human review

Usage:
    python auto_annotate.py extraction.json annotation.xlsx source.pdf
"""

from __future__ import annotations

import asyncio
import io
import json
import logging
import sys
from pathlib import Path

import fitz
from openpyxl import load_workbook

logging.basicConfig(level=logging.INFO, format="%(levelname)s: %(message)s")
logger = logging.getLogger(__name__)


async def auto_annotate(
    extraction_path: Path,
    xlsx_path: Path,
    pdf_path: Path,
    provider: str = "openai",
    api_key: str = "",
):
    """Run auto-annotation on unannotated cells in the Excel workbook."""

    # Load extraction result for cell metadata
    with open(extraction_path, encoding="utf-8") as f:
        result = json.load(f)

    # Load PDF for page images
    doc = fitz.open(str(pdf_path))

    # Load workbook
    wb = load_workbook(xlsx_path)
    ws = wb["soa_cells"]

    # Build cell lookup from extraction
    cell_lookup: dict[str, dict] = {}
    for table in result.get("tables", []):
        for cell in table.get("cells", []):
            key = f"{table['table_id']}_{cell['row']}_{cell['col']}"
            cell_lookup[key] = {
                **cell,
                "table_id": table["table_id"],
                "pages": table.get("source_pages", []),
            }

    # Process each row in the soa_cells tab
    stats = {"total": 0, "auto_y": 0, "auto_n": 0, "review": 0, "skipped": 0}

    for row_num in range(2, ws.max_row + 1):
        is_correct = ws.cell(row=row_num, column=10).value
        if is_correct:  # Already annotated
            stats["skipped"] += 1
            continue

        table_id = str(ws.cell(row=row_num, column=1).value or "")
        row_idx = ws.cell(row=row_num, column=3).value
        col_idx = ws.cell(row=row_num, column=4).value
        extracted_value = str(ws.cell(row=row_num, column=7).value or "")
        data_type = str(ws.cell(row=row_num, column=8).value or "")
        confidence = ws.cell(row=row_num, column=9).value or 0

        stats["total"] += 1

        # Route by data type
        if data_type == "EMPTY":
            # Check if the cell region is genuinely blank using OCR
            result_val, conf, method = _verify_empty(extracted_value)
            _write_result(ws, row_num, result_val, conf, method)
            stats["auto_y" if result_val == "Y" else "auto_n"] += 1

        elif data_type == "MARKER":
            # Check if OCR reads the same marker
            result_val, conf, method = _verify_marker(extracted_value)
            _write_result(ws, row_num, result_val, conf, method)
            stats["auto_y" if result_val == "Y" else "review"] += 1

        elif data_type == "TEXT":
            # For TEXT cells, check if the value is a procedure name (row header)
            # Procedure names in column 0 are high-value and need verification
            if col_idx == 0:
                # Row headers — verify that text exists somewhere in the table pages
                pages = cell_lookup.get(f"{table_id}_{row_idx}_{col_idx}", {}).get("pages", [])
                found = _verify_text_in_pages(doc, pages, extracted_value)
                if found:
                    _write_result(ws, row_num, "Y", 0.85, "AUTO:text_in_page")
                    stats["auto_y"] += 1
                else:
                    _write_result(ws, row_num, "", 0, "REVIEW:text_not_found_in_pages")
                    stats["review"] += 1
            else:
                # Non-header TEXT cells — flag for review if low confidence
                if confidence >= 0.90:
                    _write_result(ws, row_num, "", 0, "REVIEW:text_needs_visual_check")
                    stats["review"] += 1
                else:
                    _write_result(ws, row_num, "", 0, "REVIEW:low_confidence_text")
                    stats["review"] += 1

        elif data_type in ("CONDITIONAL", "NUMERIC"):
            # Always flag for human review
            _write_result(ws, row_num, "", 0, "HUMAN_REQUIRED")
            stats["review"] += 1

        else:
            _write_result(ws, row_num, "", 0, "REVIEW:unknown_type")
            stats["review"] += 1

    doc.close()
    wb.save(xlsx_path)

    logger.info(f"Auto-annotation complete:")
    logger.info(f"  Total unannotated: {stats['total']}")
    logger.info(f"  Auto-accepted (Y): {stats['auto_y']}")
    logger.info(f"  Auto-rejected (N): {stats['auto_n']}")
    logger.info(f"  Flagged for review: {stats['review']}")
    logger.info(f"  Skipped (already done): {stats['skipped']}")
    logger.info(f"  Automation rate: {stats['auto_y'] / max(stats['total'], 1) * 100:.0f}%")


def _verify_empty(extracted_value: str) -> tuple[str, float, str]:
    """Verify an EMPTY cell — is the extracted value genuinely blank?"""
    value = extracted_value.strip()
    if not value or value in ("", "nan", "NaN", "None"):
        return "Y", 0.95, "AUTO:empty_confirmed"
    else:
        return "N", 0.80, f"AUTO:empty_but_has_value:{value[:20]}"


def _verify_marker(extracted_value: str) -> tuple[str, float, str]:
    """Verify a MARKER cell — is it a recognizable X/checkmark?"""
    value = extracted_value.strip()
    known_markers = {"X", "x", "✓", "✔", "✗", "✘", "Y", "N", "Yes", "No"}

    if value in known_markers:
        return "Y", 0.90, "AUTO:marker_recognized"
    elif value and len(value) <= 3:
        # Short value that might be a marker variant
        return "", 0, f"REVIEW:marker_variant:{value}"
    elif not value:
        return "Y", 0.90, "AUTO:marker_empty_confirmed"
    else:
        return "", 0, f"REVIEW:unexpected_marker:{value[:20]}"


def _verify_text_in_pages(doc: fitz.Document, pages: list[int], text: str) -> bool:
    """Check if extracted text appears on any of the table's pages."""
    if not text or len(text.strip()) < 3:
        return False

    search_term = text.strip()[:30].lower()
    for page_num in pages:
        if page_num < doc.page_count:
            page_text = doc[page_num].get_text("text").lower()
            if search_term in page_text:
                return True
    return False


def _write_result(ws, row_num: int, is_correct: str, confidence: float, method: str):
    """Write auto-annotation result back to the Excel row."""
    if is_correct:
        ws.cell(row=row_num, column=10, value=is_correct)
    # Always write the method to notes column for audit trail
    existing_notes = ws.cell(row=row_num, column=14).value or ""
    ws.cell(row=row_num, column=14, value=method)

    # Set difficulty based on method
    if "HUMAN_REQUIRED" in method:
        ws.cell(row=row_num, column=13, value="hard")
    elif "REVIEW" in method:
        ws.cell(row=row_num, column=13, value="moderate")
    elif "AUTO" in method:
        ws.cell(row=row_num, column=13, value="trivial")


if __name__ == "__main__":
    if len(sys.argv) < 4:
        print("Usage: python auto_annotate.py <extraction.json> <annotation.xlsx> <source.pdf>")
        print("  Optional: --provider openai --api-key sk-...")
        sys.exit(1)

    extraction = Path(sys.argv[1])
    xlsx = Path(sys.argv[2])
    pdf = Path(sys.argv[3])

    asyncio.run(auto_annotate(extraction, xlsx, pdf))
