"""
Convert completed annotation spreadsheet to OmniDocBench-compatible JSON.

Reads the annotated Excel workbook and produces ground truth JSON suitable
for TEDS computation and cell-level accuracy measurement.

Usage:
    python convert_to_json.py annotation.xlsx output.json
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import load_workbook


def convert_annotation_to_json(xlsx_path: Path, output_path: Path):
    wb = load_workbook(xlsx_path, read_only=True)
    result = {"version": "1.0", "source": str(xlsx_path), "tables": [], "footnotes": [], "procedures": []}

    # Tab 1: soa_cells
    if "soa_cells" in wb.sheetnames:
        ws = wb["soa_cells"]
        tables: dict[str, dict] = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue
            table_id = str(row[0])
            if table_id not in tables:
                tables[table_id] = {
                    "table_id": table_id,
                    "source_pages": row[1],
                    "ground_truth_cells": [],
                    "stats": {"total": 0, "correct": 0, "wrong": 0, "unchecked": 0},
                }

            is_correct = str(row[9] or "").strip().upper()
            extracted = str(row[6] or "")
            correct_val = str(row[10] or "") if is_correct == "N" else extracted

            cell = {
                "row": int(row[2] or 0),
                "col": int(row[3] or 0),
                "value": correct_val if is_correct == "N" else extracted,
                "extracted_value": extracted,
                "data_type": str(row[7] or "TEXT"),
                "is_correct": is_correct == "Y",
                "confidence": float(row[8] or 0),
                "footnote_markers": [m.strip() for m in str(row[11] or "").split(",") if m.strip()],
                "difficulty": str(row[12] or "moderate"),
            }

            tables[table_id]["ground_truth_cells"].append(cell)
            tables[table_id]["stats"]["total"] += 1
            if is_correct == "Y":
                tables[table_id]["stats"]["correct"] += 1
            elif is_correct == "N":
                tables[table_id]["stats"]["wrong"] += 1
            else:
                tables[table_id]["stats"]["unchecked"] += 1

        result["tables"] = list(tables.values())

    # Tab 2: footnotes
    if "footnotes" in wb.sheetnames:
        ws = wb["footnotes"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            is_correct = str(row[6] or "").strip().upper()
            bindings_str = str(row[4] or "")
            applies_to = []
            for pair in bindings_str.split(","):
                pair = pair.strip()
                if ":" in pair:
                    parts = pair.split(":")
                    try:
                        applies_to.append({"row": int(parts[0]), "col": int(parts[1])})
                    except ValueError:
                        pass

            fn = {
                "table_id": str(row[0]),
                "marker": str(row[1] or ""),
                "text": str(row[7] or row[2] or ""),  # correct_text or extracted
                "footnote_type": str(row[8] or row[3] or ""),  # correct_type or extracted
                "applies_to": applies_to,
                "is_correct": is_correct == "Y",
            }
            result["footnotes"].append(fn)

    # Tab 3: procedures
    if "procedures" in wb.sheetnames:
        ws = wb["procedures"]
        for row in ws.iter_rows(min_row=2, values_only=True):
            if not row or not row[0]:
                continue

            is_correct = str(row[6] or "").strip().upper()
            proc = {
                "raw_name": str(row[0] or ""),
                "canonical_name": str(row[7] or row[1] or ""),
                "cpt_code": str(row[8] or row[2] or ""),
                "category": str(row[4] or ""),
                "cost_tier": str(row[5] or ""),
                "is_correct": is_correct == "Y",
            }
            result["procedures"].append(proc)

    # Summary stats
    total_cells = sum(t["stats"]["total"] for t in result["tables"])
    correct_cells = sum(t["stats"]["correct"] for t in result["tables"])
    wrong_cells = sum(t["stats"]["wrong"] for t in result["tables"])
    unchecked = sum(t["stats"]["unchecked"] for t in result["tables"])

    result["summary"] = {
        "total_cells": total_cells,
        "correct_cells": correct_cells,
        "wrong_cells": wrong_cells,
        "unchecked_cells": unchecked,
        "cell_accuracy": correct_cells / max(total_cells - unchecked, 1),
        "footnotes_annotated": len(result["footnotes"]),
        "procedures_annotated": len(result["procedures"]),
    }

    with open(output_path, "w", encoding="utf-8") as f:
        json.dump(result, f, indent=2, ensure_ascii=False)

    print(f"Ground truth JSON saved: {output_path}")
    print(f"  Cells: {total_cells} total, {correct_cells} correct, {wrong_cells} wrong, {unchecked} unchecked")
    print(f"  Accuracy: {result['summary']['cell_accuracy']:.1%}")
    print(f"  Footnotes: {len(result['footnotes'])}")
    print(f"  Procedures: {len(result['procedures'])}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python convert_to_json.py <annotation.xlsx> <output.json>")
        sys.exit(1)

    convert_annotation_to_json(Path(sys.argv[1]), Path(sys.argv[2]))
