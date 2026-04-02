"""
Excel Ingestor — converts an XLSX file to FormattedDocument IR.

Uses openpyxl to parse worksheets, extracting cell values with formatting
metadata (bold, italic, font, size, color). Each worksheet maps to a
FormattedPage containing a single FormattedTable. Merged cells are
represented via rowspan/colspan on FormattedTableCell.
"""

from __future__ import annotations

import io
import logging
from typing import Any

from src.formatter.extractor import (
    FormattedDocument,
    FormattedLine,
    FormattedPage,
    FormattedParagraph,
    FormattedSpan,
    FormattedTable,
    FormattedTableCell,
)

logger = logging.getLogger(__name__)

try:
    import openpyxl
    from openpyxl.cell.cell import Cell
    from openpyxl.utils import get_column_letter

    _HAS_OPENPYXL = True
except ImportError:
    _HAS_OPENPYXL = False

# Default page dimensions for generated pages (US Letter in points)
_DEFAULT_WIDTH = 612.0
_DEFAULT_HEIGHT = 792.0


class ExcelIngestor:
    """Parses an XLSX file into a FormattedDocument IR.

    Each worksheet becomes a FormattedPage with a single FormattedTable.

    Usage::

        ingestor = ExcelIngestor()
        doc = ingestor.ingest(xlsx_bytes, filename="data.xlsx")
    """

    def ingest(self, content: bytes, filename: str = "") -> FormattedDocument:
        """Parse an XLSX file and return a FormattedDocument.

        Args:
            content: Raw bytes of the .xlsx file.
            filename: Original filename for metadata.

        Returns:
            A FormattedDocument populated from the workbook.

        Raises:
            ImportError: If openpyxl is not installed.
        """
        if not _HAS_OPENPYXL:
            raise ImportError(
                "openpyxl is required for Excel ingestion. "
                "Install it with: pip install openpyxl"
            )

        if not content:
            return FormattedDocument(filename=filename)

        wb = openpyxl.load_workbook(
            io.BytesIO(content),
            read_only=False,
            data_only=True,
        )

        pages: list[FormattedPage] = []
        font_counts: dict[str, int] = {}
        color_counts: dict[str, int] = {}

        for sheet_idx, ws in enumerate(wb.worksheets):
            page = self._extract_sheet(ws, sheet_idx, font_counts, color_counts)
            pages.append(page)

        wb.close()

        # Build style inventory (all tables)
        style_counts: dict[str, int] = {"table_cell": 0}
        for page in pages:
            for table in page.tables:
                for row in table.rows:
                    for cell in row:
                        style_counts["table_cell"] += 1

        return FormattedDocument(
            filename=filename,
            pages=pages,
            font_inventory=font_counts,
            color_inventory=color_counts,
            style_inventory=style_counts,
        )

    # ------------------------------------------------------------------
    # Internal — sheet extraction
    # ------------------------------------------------------------------

    def _extract_sheet(
        self,
        ws: Any,
        sheet_idx: int,
        font_counts: dict[str, int],
        color_counts: dict[str, int],
    ) -> FormattedPage:
        """Extract a FormattedPage from a single worksheet."""
        page = FormattedPage(
            page_number=sheet_idx,
            width=_DEFAULT_WIDTH,
            height=_DEFAULT_HEIGHT,
        )

        # Build merged-cell lookup: {(row, col): (rowspan, colspan)}
        merge_map = self._build_merge_map(ws)

        # Determine table dimensions
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        if max_row == 0 or max_col == 0:
            return page

        rows: list[list[FormattedTableCell]] = []

        for r_idx in range(1, max_row + 1):
            row_cells: list[FormattedTableCell] = []

            for c_idx in range(1, max_col + 1):
                cell = ws.cell(row=r_idx, column=c_idx)

                # Get text value
                text = self._cell_text(cell)

                # Formatting
                is_bold = False
                is_italic = False
                font_name = ""
                font_size = 11.0
                color_int = 0

                try:
                    cell_font = cell.font
                    if cell_font is not None:
                        is_bold = bool(cell_font.bold)
                        is_italic = bool(cell_font.italic)
                        font_name = cell_font.name or ""
                        if cell_font.size is not None:
                            font_size = float(cell_font.size)
                        color_int = _extract_font_color(cell_font)
                except Exception:
                    pass

                # Merged cell spans
                rowspan = 1
                colspan = 1
                merge_key = (r_idx, c_idx)
                if merge_key in merge_map:
                    rowspan, colspan = merge_map[merge_key]

                # First row is typically a header
                is_header = r_idx == 1

                row_cells.append(FormattedTableCell(
                    text=text,
                    row=r_idx - 1,  # 0-indexed
                    col=c_idx - 1,
                    rowspan=rowspan,
                    colspan=colspan,
                    bold=is_bold or is_header,
                    is_header=is_header,
                ))

                # Track inventories
                if font_name:
                    font_counts[font_name] = font_counts.get(font_name, 0) + 1
                hex_color = f"#{color_int:06X}"
                color_counts[hex_color] = color_counts.get(hex_color, 0) + 1

            rows.append(row_cells)

        table = FormattedTable(
            rows=rows,
            num_rows=max_row,
            num_cols=max_col,
            y_position=0.0,
        )

        page.tables.append(table)

        # Add sheet title as a heading paragraph if available
        title = ws.title
        if title:
            title_span = FormattedSpan(
                text=title,
                x0=0, y0=0, x1=200, y1=14,
                font="Arial",
                size=14.0,
                bold=True,
            )
            title_line = FormattedLine(spans=[title_span], y_center=0.0, indent=0.0)
            title_para = FormattedParagraph(
                lines=[title_line],
                style="heading2",
                spacing_after=8.0,
            )
            page.paragraphs.append(title_para)

        return page

    # ------------------------------------------------------------------
    # Internal — merged cells
    # ------------------------------------------------------------------

    def _build_merge_map(self, ws: Any) -> dict[tuple[int, int], tuple[int, int]]:
        """Build a lookup of top-left cells of merged ranges to (rowspan, colspan)."""
        merge_map: dict[tuple[int, int], tuple[int, int]] = {}

        try:
            for merged_range in ws.merged_cells.ranges:
                min_row = merged_range.min_row
                min_col = merged_range.min_col
                max_row = merged_range.max_row
                max_col = merged_range.max_col

                rowspan = max_row - min_row + 1
                colspan = max_col - min_col + 1

                merge_map[(min_row, min_col)] = (rowspan, colspan)
        except Exception as exc:
            logger.debug("Failed to parse merged cells: %s", exc)

        return merge_map

    # ------------------------------------------------------------------
    # Internal — helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _cell_text(cell: Any) -> str:
        """Extract text from an openpyxl cell, handling None and types."""
        value = cell.value
        if value is None:
            return ""
        if isinstance(value, (int, float)):
            # Avoid trailing .0 for whole numbers
            if isinstance(value, float) and value == int(value):
                return str(int(value))
            return str(value)
        return str(value).strip()


# ---------------------------------------------------------------------------
# Module-level helpers
# ---------------------------------------------------------------------------

def _extract_font_color(font: Any) -> int:
    """Extract a packed RGB integer from an openpyxl Font's color.

    Handles theme colors gracefully by returning 0 (black) when the color
    is theme-based or otherwise unavailable.
    """
    try:
        if font.color is None:
            return 0
        color = font.color

        # Direct RGB value
        if color.rgb and color.rgb != "00000000":
            rgb_str = str(color.rgb)
            # openpyxl may return ARGB as "FFRRGGBB"
            if len(rgb_str) == 8:
                rgb_str = rgb_str[2:]  # strip alpha
            if len(rgb_str) == 6:
                return int(rgb_str, 16)

        # Theme color — cannot reliably resolve without theme XML
        if color.theme is not None:
            return 0

    except Exception:
        pass

    return 0
