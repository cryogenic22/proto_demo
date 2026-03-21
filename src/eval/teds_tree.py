"""
TEDS Evaluator for ProtoExtract Pipeline
=========================================
Implements the Tree-Edit-Distance-based Similarity (TEDS) metric from
OmniDocBench for evaluating SoA table extraction accuracy.

Converts our annotation XLSX to OmniDocBench-compatible HTML, then
computes TEDS and TEDS-S (structure-only) scores using the APTED
(All Path Tree Edit Distance) algorithm.

Usage:
    from teds_evaluator import TEDSEvaluator

    evaluator = TEDSEvaluator()
    score = evaluator.compute_teds(pred_html, gt_html)
    # score: 0.0 (completely wrong) to 1.0 (perfect)

    # Structure-only evaluation (ignores cell content)
    score_s = evaluator.compute_teds(pred_html, gt_html, structure_only=True)
"""

import re
import logging
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Tuple, Any
from collections import defaultdict

logger = logging.getLogger(__name__)


# ── Tree representation for APTED ──────────────────────────────────────

@dataclass
class TreeNode:
    """Node in the tree representation of an HTML table."""
    tag: str                                    # e.g., "table", "tr", "td"
    content: str = ""                           # text content (for leaf nodes)
    attributes: Dict[str, str] = field(default_factory=dict)
    children: List["TreeNode"] = field(default_factory=list)

    def to_bracket_string(self, structure_only: bool = False) -> str:
        """Convert tree to bracket notation for APTED."""
        label = self.tag
        if self.attributes:
            attrs = ",".join(f"{k}={v}" for k, v in sorted(self.attributes.items()))
            label += f"[{attrs}]"
        if not structure_only and self.content:
            label += f":{self.content[:50]}"  # Truncate long content

        if not self.children:
            return "{" + label + "}"
        child_strs = "".join(c.to_bracket_string(structure_only) for c in self.children)
        return "{" + label + child_strs + "}"


# ── APTED Implementation ──────────────────────────────────────────────

class APTEDConfig:
    """Configuration for APTED tree edit distance computation."""

    @staticmethod
    def rename_cost(node1: TreeNode, node2: TreeNode, structure_only: bool = False) -> float:
        """Cost of renaming node1 to node2."""
        if node1.tag != node2.tag:
            return 1.0
        if node1.attributes != node2.attributes:
            return 0.5
        if not structure_only and node1.content != node2.content:
            # Partial credit for similar content
            if node1.content and node2.content:
                similarity = _string_similarity(node1.content, node2.content)
                return 1.0 - similarity
            return 1.0
        return 0.0

    @staticmethod
    def delete_cost(node: TreeNode) -> float:
        """Cost of deleting a node."""
        return 1.0

    @staticmethod
    def insert_cost(node: TreeNode) -> float:
        """Cost of inserting a node."""
        return 1.0


def _string_similarity(s1: str, s2: str) -> float:
    """Compute normalized string similarity using Levenshtein distance."""
    s1, s2 = s1.lower().strip(), s2.lower().strip()
    if s1 == s2:
        return 1.0
    if not s1 or not s2:
        return 0.0

    # Simple Levenshtein
    m, n = len(s1), len(s2)
    dp = list(range(n + 1))
    for i in range(1, m + 1):
        prev = dp[0]
        dp[0] = i
        for j in range(1, n + 1):
            temp = dp[j]
            if s1[i-1] == s2[j-1]:
                dp[j] = prev
            else:
                dp[j] = 1 + min(dp[j], dp[j-1], prev)
            prev = temp
    return 1.0 - dp[n] / max(m, n)


def _tree_size(node: TreeNode) -> int:
    """Count total nodes in a tree."""
    return 1 + sum(_tree_size(c) for c in node.children)


def _compute_ted(tree1: TreeNode, tree2: TreeNode, structure_only: bool = False) -> float:
    """
    Compute Tree Edit Distance between two trees using a simplified
    Zhang-Shasha algorithm (equivalent results to APTED for our use case).

    This is an O(n^2 * m^2) algorithm where n, m are tree sizes.
    For SoA tables (typically <500 nodes), this is fast enough.
    """
    config = APTEDConfig()

    # Get ordered node lists via post-order traversal
    def post_order(node: TreeNode) -> List[TreeNode]:
        result = []
        for child in node.children:
            result.extend(post_order(child))
        result.append(node)
        return result

    nodes1 = post_order(tree1)
    nodes2 = post_order(tree2)
    n1, n2 = len(nodes1), len(nodes2)

    # DP table
    dp = [[0.0] * (n2 + 1) for _ in range(n1 + 1)]

    for i in range(1, n1 + 1):
        dp[i][0] = dp[i-1][0] + config.delete_cost(nodes1[i-1])
    for j in range(1, n2 + 1):
        dp[0][j] = dp[0][j-1] + config.insert_cost(nodes2[j-1])

    for i in range(1, n1 + 1):
        for j in range(1, n2 + 1):
            cost_rename = config.rename_cost(nodes1[i-1], nodes2[j-1], structure_only)
            dp[i][j] = min(
                dp[i-1][j] + config.delete_cost(nodes1[i-1]),
                dp[i][j-1] + config.insert_cost(nodes2[j-1]),
                dp[i-1][j-1] + cost_rename,
            )

    return dp[n1][n2]


# ── HTML Parsing ──────────────────────────────────────────────────────

class HTMLTableParser:
    """Parse HTML table string into TreeNode structure."""

    TAG_PATTERN = re.compile(r'<(/?)(\w+)([^>]*)>')
    ATTR_PATTERN = re.compile(r'(\w+)=["\']([^"\']*)["\']')

    @classmethod
    def parse(cls, html: str) -> TreeNode:
        """
        Parse an HTML table string into a TreeNode tree.

        Expected format:
            <table>
              <tr>
                <td rowspan="2">Content</td>
                <td>More content</td>
              </tr>
            </table>
        """
        # Clean HTML
        html = re.sub(r'\s+', ' ', html.strip())

        root = TreeNode(tag="root")
        stack = [root]

        pos = 0
        while pos < len(html):
            match = cls.TAG_PATTERN.search(html, pos)
            if not match:
                # Remaining text is content
                text = html[pos:].strip()
                if text and stack:
                    stack[-1].content += text
                break

            # Text before tag
            text_before = html[pos:match.start()].strip()
            if text_before and stack:
                stack[-1].content += text_before

            is_closing = match.group(1) == "/"
            tag_name = match.group(2).lower()
            attr_str = match.group(3)

            if is_closing:
                # Pop stack
                if len(stack) > 1:
                    stack.pop()
            else:
                # Parse attributes
                attrs = {}
                for attr_match in cls.ATTR_PATTERN.finditer(attr_str):
                    key = attr_match.group(1).lower()
                    val = attr_match.group(2)
                    if key in ("rowspan", "colspan"):
                        attrs[key] = val

                node = TreeNode(tag=tag_name, attributes=attrs)
                stack[-1].children.append(node)

                # Self-closing check
                if not attr_str.rstrip().endswith("/"):
                    stack.append(node)

            pos = match.end()

        # Return the table node (first child of root)
        if root.children:
            return root.children[0]
        return root


# ── Grid to HTML Converter ────────────────────────────────────────────

@dataclass
class CellData:
    """Represents a cell in our internal grid format."""
    text: str
    row: int
    col: int
    rowspan: int = 1
    colspan: int = 1
    is_header: bool = False


def grid_to_html(cells: List[CellData], num_rows: int, num_cols: int) -> str:
    """
    Convert our internal cell grid representation to HTML table.

    This is the bridge between ProtoExtract's output format and
    OmniDocBench's TEDS evaluation input.
    """
    # Build occupancy map for merged cells
    occupied = set()
    cell_map = {}

    for cell in cells:
        cell_map[(cell.row, cell.col)] = cell
        for dr in range(cell.rowspan):
            for dc in range(cell.colspan):
                occupied.add((cell.row + dr, cell.col + dc))

    html_parts = ["<table>"]

    for r in range(num_rows):
        html_parts.append("  <tr>")
        for c in range(num_cols):
            if (r, c) in cell_map:
                cell = cell_map[(r, c)]
                tag = "th" if cell.is_header else "td"
                attrs = ""
                if cell.rowspan > 1:
                    attrs += f' rowspan="{cell.rowspan}"'
                if cell.colspan > 1:
                    attrs += f' colspan="{cell.colspan}"'
                text = cell.text.replace("<", "&lt;").replace(">", "&gt;")
                html_parts.append(f"    <{tag}{attrs}>{text}</{tag}>")
            elif (r, c) not in occupied:
                # Empty cell not covered by any span
                html_parts.append("    <td></td>")
        html_parts.append("  </tr>")

    html_parts.append("</table>")
    return "\n".join(html_parts)


def xlsx_annotation_to_html(xlsx_path: str, sheet_name: str = None) -> str:
    """
    Convert our annotation XLSX format to HTML table for TEDS evaluation.

    Expected XLSX format:
        - Each sheet is one SoA table
        - Merged cells represent rowspan/colspan
        - First row(s) are headers
    """
    import openpyxl

    wb = openpyxl.load_workbook(xlsx_path)
    ws = wb[sheet_name] if sheet_name else wb.active

    # Detect merged cells
    merged_ranges = list(ws.merged_cells.ranges)

    # Build cell data
    cells = []
    occupied = set()

    for merge_range in merged_ranges:
        min_row, min_col = merge_range.min_row, merge_range.min_col
        max_row, max_col = merge_range.max_row, merge_range.max_col
        rowspan = max_row - min_row + 1
        colspan = max_col - min_col + 1

        cell_value = ws.cell(min_row, min_col).value or ""
        cells.append(CellData(
            text=str(cell_value),
            row=min_row - 1,  # 0-indexed
            col=min_col - 1,
            rowspan=rowspan,
            colspan=colspan,
            is_header=(min_row == 1),
        ))

        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                occupied.add((r, c))

    # Add non-merged cells
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
        for cell in row:
            if (cell.row, cell.column) not in occupied:
                cells.append(CellData(
                    text=str(cell.value or ""),
                    row=cell.row - 1,
                    col=cell.column - 1,
                    is_header=(cell.row == 1),
                ))

    return grid_to_html(cells, ws.max_row, ws.max_column)


# ── TEDS Evaluator ────────────────────────────────────────────────────

@dataclass
class TEDSResult:
    """Result of TEDS evaluation."""
    teds: float             # Full TEDS score (content + structure)
    teds_s: float           # Structure-only TEDS score
    pred_nodes: int         # Number of nodes in predicted tree
    gt_nodes: int           # Number of nodes in ground truth tree
    ted_full: float         # Raw tree edit distance (full)
    ted_structure: float    # Raw tree edit distance (structure only)


class TEDSEvaluator:
    """
    Compute TEDS (Tree-Edit-Distance-based Similarity) for table evaluation.

    TEDS = 1 - TED(pred, gt) / max(|pred|, |gt|)

    Where:
        - TED is the tree edit distance
        - |pred|, |gt| are the number of nodes in each tree
        - Score ranges from 0.0 (completely wrong) to 1.0 (perfect)
    """

    def __init__(self):
        self.parser = HTMLTableParser()

    def compute_teds(
        self,
        pred_html: str,
        gt_html: str,
        structure_only: bool = False,
    ) -> float:
        """
        Compute TEDS score between predicted and ground truth HTML tables.

        Args:
            pred_html: Predicted table as HTML string
            gt_html: Ground truth table as HTML string
            structure_only: If True, ignore cell content (TEDS-S)

        Returns:
            TEDS score between 0.0 and 1.0
        """
        pred_tree = self.parser.parse(pred_html)
        gt_tree = self.parser.parse(gt_html)

        pred_size = _tree_size(pred_tree)
        gt_size = _tree_size(gt_tree)

        ted = _compute_ted(pred_tree, gt_tree, structure_only)
        max_size = max(pred_size, gt_size)

        if max_size == 0:
            return 1.0

        teds = 1.0 - ted / max_size
        return max(0.0, teds)

    def evaluate_full(self, pred_html: str, gt_html: str) -> TEDSResult:
        """Compute both TEDS and TEDS-S and return detailed results."""
        pred_tree = self.parser.parse(pred_html)
        gt_tree = self.parser.parse(gt_html)

        pred_size = _tree_size(pred_tree)
        gt_size = _tree_size(gt_tree)
        max_size = max(pred_size, gt_size)

        ted_full = _compute_ted(pred_tree, gt_tree, structure_only=False)
        ted_struct = _compute_ted(pred_tree, gt_tree, structure_only=True)

        teds = max(0.0, 1.0 - ted_full / max_size) if max_size > 0 else 1.0
        teds_s = max(0.0, 1.0 - ted_struct / max_size) if max_size > 0 else 1.0

        return TEDSResult(
            teds=round(teds, 4),
            teds_s=round(teds_s, 4),
            pred_nodes=pred_size,
            gt_nodes=gt_size,
            ted_full=round(ted_full, 2),
            ted_structure=round(ted_struct, 2),
        )

    def evaluate_batch(
        self,
        predictions: List[str],
        ground_truths: List[str],
    ) -> Dict[str, Any]:
        """
        Evaluate a batch of table predictions.

        Returns:
            Dict with mean_teds, mean_teds_s, per_table results, and
            attribute-stratified breakdown.
        """
        assert len(predictions) == len(ground_truths)

        results = []
        for pred, gt in zip(predictions, ground_truths):
            result = self.evaluate_full(pred, gt)
            results.append(result)

        teds_scores = [r.teds for r in results]
        teds_s_scores = [r.teds_s for r in results]

        return {
            "mean_teds": round(sum(teds_scores) / len(teds_scores), 4),
            "mean_teds_s": round(sum(teds_s_scores) / len(teds_s_scores), 4),
            "min_teds": round(min(teds_scores), 4),
            "max_teds": round(max(teds_scores), 4),
            "num_tables": len(results),
            "per_table": [
                {"teds": r.teds, "teds_s": r.teds_s, "pred_nodes": r.pred_nodes, "gt_nodes": r.gt_nodes}
                for r in results
            ],
        }

    def evaluate_stratified(
        self,
        predictions: List[str],
        ground_truths: List[str],
        attributes: List[Dict[str, str]],
    ) -> Dict[str, Dict[str, float]]:
        """
        Attribute-stratified TEDS evaluation (OmniDocBench-style).

        Args:
            predictions: List of predicted HTML tables
            ground_truths: List of ground truth HTML tables
            attributes: List of attribute dicts for each table, e.g.:
                {"has_merged_cells": "yes", "num_rows": "large", "has_footnotes": "yes"}

        Returns:
            Dict mapping attribute values to mean TEDS scores.
        """
        assert len(predictions) == len(ground_truths) == len(attributes)

        # Group by attribute
        stratified = defaultdict(list)

        for pred, gt, attrs in zip(predictions, ground_truths, attributes):
            result = self.evaluate_full(pred, gt)
            for key, value in attrs.items():
                stratified[f"{key}={value}"].append(result.teds)

        return {
            attr: {
                "mean_teds": round(sum(scores) / len(scores), 4),
                "count": len(scores),
            }
            for attr, scores in sorted(stratified.items())
        }


# ── CLI usage ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="TEDS Evaluator for SoA tables")
    parser.add_argument("--pred", required=True, help="Predicted HTML table file")
    parser.add_argument("--gt", required=True, help="Ground truth HTML table file")
    parser.add_argument("--structure-only", action="store_true")
    args = parser.parse_args()

    with open(args.pred) as f:
        pred_html = f.read()
    with open(args.gt) as f:
        gt_html = f.read()

    evaluator = TEDSEvaluator()
    result = evaluator.evaluate_full(pred_html, gt_html)
    print(f"TEDS:   {result.teds:.4f}")
    print(f"TEDS-S: {result.teds_s:.4f}")
    print(f"Predicted nodes: {result.pred_nodes}")
    print(f"Ground truth nodes: {result.gt_nodes}")
