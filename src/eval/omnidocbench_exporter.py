"""
OmniDocBench Exporter for ProtoExtract Pipeline
================================================
Converts our annotation XLSX to OmniDocBench JSON format, enabling
evaluation using the OmniDocBench evaluation suite on our SoA data.

OmniDocBench JSON Schema:
    {
        "dataset_name": "protoextract_soa",
        "version": "1.0",
        "tables": [
            {
                "id": "P-01_SoA_1",
                "source": "P-01.pdf",
                "page": 45,
                "html": "<table>...</table>",
                "attributes": {
                    "language": "en",
                    "has_merged_cells": true,
                    "num_rows": 25,
                    "num_cols": 15,
                    "has_footnotes": true,
                    "has_conditional_markers": true,
                    "line_density": "high",
                    "table_type": "soa"
                }
            }
        ]
    }

Usage:
    from omnidocbench_exporter import OmniDocBenchExporter

    exporter = OmniDocBenchExporter()
    exporter.add_table_from_xlsx("annotations/P-01_SoA.xlsx", "P-01.pdf", page=45)
    exporter.export("omnidocbench_eval.json")
"""

import json
import re
import logging
from dataclasses import dataclass, field
from typing import List, Dict, Optional, Any
from pathlib import Path

logger = logging.getLogger(__name__)


@dataclass
class TableAttributes:
    """OmniDocBench-compatible table attributes for stratified evaluation."""
    language: str = "en"
    has_merged_cells: bool = False
    num_rows: int = 0
    num_cols: int = 0
    has_footnotes: bool = False
    has_conditional_markers: bool = False
    line_density: str = "medium"     # low, medium, high
    table_type: str = "soa"          # soa, appendix, summary
    num_pages: int = 1               # pages the table spans
    therapeutic_area: str = ""
    sponsor: str = ""

    def to_dict(self) -> Dict[str, Any]:
        return {
            "language": self.language,
            "has_merged_cells": self.has_merged_cells,
            "num_rows": self.num_rows,
            "num_cols": self.num_cols,
            "has_footnotes": self.has_footnotes,
            "has_conditional_markers": self.has_conditional_markers,
            "line_density": self.line_density,
            "table_type": self.table_type,
            "num_pages": self.num_pages,
            "therapeutic_area": self.therapeutic_area,
            "sponsor": self.sponsor,
        }

    def to_stratification_dict(self) -> Dict[str, str]:
        """Convert to string-valued dict for TEDS stratification."""
        return {
            "language": self.language,
            "has_merged_cells": "yes" if self.has_merged_cells else "no",
            "size": "large" if self.num_rows > 20 else ("medium" if self.num_rows > 10 else "small"),
            "has_footnotes": "yes" if self.has_footnotes else "no",
            "line_density": self.line_density,
            "multi_page": "yes" if self.num_pages > 1 else "no",
        }


@dataclass
class TableEntry:
    """A single table entry in OmniDocBench format."""
    id: str
    source: str
    page: int
    html: str
    attributes: TableAttributes = field(default_factory=TableAttributes)

    def to_dict(self) -> Dict[str, Any]:
        return {
            "id": self.id,
            "source": self.source,
            "page": self.page,
            "html": self.html,
            "attributes": self.attributes.to_dict(),
        }


class OmniDocBenchExporter:
    """
    Converts ProtoExtract annotations to OmniDocBench evaluation format.

    Workflow:
        1. Add tables from XLSX annotations or HTML strings
        2. Auto-detect table attributes (merged cells, footnotes, etc.)
        3. Export to OmniDocBench JSON
        4. Run OmniDocBench evaluation suite
    """

    def __init__(self, dataset_name: str = "protoextract_soa", version: str = "1.0"):
        self.dataset_name = dataset_name
        self.version = version
        self.tables: List[TableEntry] = []

    # ── Conditional marker patterns common in SoA tables ──────────────

    CONDITIONAL_MARKERS = {
        "if applicable", "as needed", "per investigator",
        "if clinically indicated", "optional", "prn",
        "at discretion", "as required", "if available",
    }

    FOOTNOTE_PATTERNS = [
        r'[a-z]\)',          # a), b), c)
        r'\d+\)',            # 1), 2), 3)
        r'[*†‡§¶#]',        # symbols
        r'Note:',            # explicit notes
        r'Footnote',
    ]

    def _detect_attributes(self, html: str) -> TableAttributes:
        """Auto-detect table attributes from HTML content."""
        attrs = TableAttributes()

        # Count rows and columns
        rows = re.findall(r'<tr[^>]*>', html)
        attrs.num_rows = len(rows)

        # Count max columns in any row
        for row_match in re.finditer(r'<tr[^>]*>(.*?)</tr>', html, re.DOTALL):
            cells = re.findall(r'<t[dh][^>]*>', row_match.group(1))
            attrs.num_cols = max(attrs.num_cols, len(cells))

        # Detect merged cells
        attrs.has_merged_cells = bool(
            re.search(r'(rowspan|colspan)=["\']([2-9]|\d{2,})', html)
        )

        # Detect footnotes
        for pattern in self.FOOTNOTE_PATTERNS:
            if re.search(pattern, html):
                attrs.has_footnotes = True
                break

        # Detect conditional markers
        html_lower = html.lower()
        attrs.has_conditional_markers = any(
            marker in html_lower for marker in self.CONDITIONAL_MARKERS
        )

        # Estimate line density
        total_cells = len(re.findall(r'<t[dh][^>]*>', html))
        if total_cells > 200:
            attrs.line_density = "high"
        elif total_cells > 50:
            attrs.line_density = "medium"
        else:
            attrs.line_density = "low"

        return attrs

    def add_table(
        self,
        table_id: str,
        source: str,
        page: int,
        html: str,
        attributes: Optional[TableAttributes] = None,
    ):
        """Add a table from an HTML string."""
        if attributes is None:
            attributes = self._detect_attributes(html)

        self.tables.append(TableEntry(
            id=table_id,
            source=source,
            page=page,
            html=html,
            attributes=attributes,
        ))
        logger.info(f"Added table {table_id} ({attributes.num_rows}x{attributes.num_cols})")

    def add_table_from_xlsx(
        self,
        xlsx_path: str,
        source_pdf: str,
        page: int = 0,
        table_id: str = None,
        sheet_name: str = None,
    ):
        """
        Add a table from our annotation XLSX format.

        Converts the XLSX to HTML and auto-detects attributes.
        """
        import openpyxl

        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb[sheet_name] if sheet_name else wb.active

        if not table_id:
            stem = Path(xlsx_path).stem
            table_id = f"{stem}_page{page}"

        # Build HTML from worksheet
        html = self._worksheet_to_html(ws)
        attributes = self._detect_attributes(html)

        # Enhance attributes from XLSX metadata
        attributes.num_rows = ws.max_row
        attributes.num_cols = ws.max_column

        self.add_table(table_id, source_pdf, page, html, attributes)
        return html

    def _worksheet_to_html(self, ws) -> str:
        """Convert an openpyxl worksheet to HTML table."""
        merged_ranges = list(ws.merged_cells.ranges)

        # Build merge map
        merge_map = {}  # (row, col) -> (rowspan, colspan)
        occupied = set()

        for merge_range in merged_ranges:
            min_r, min_c = merge_range.min_row, merge_range.min_col
            max_r, max_c = merge_range.max_row, merge_range.max_col
            merge_map[(min_r, min_c)] = (max_r - min_r + 1, max_c - min_c + 1)
            for r in range(min_r, max_r + 1):
                for c in range(min_c, max_c + 1):
                    if (r, c) != (min_r, min_c):
                        occupied.add((r, c))

        html_parts = ["<table>"]

        for row_idx in range(1, ws.max_row + 1):
            html_parts.append("  <tr>")
            for col_idx in range(1, ws.max_column + 1):
                if (row_idx, col_idx) in occupied:
                    continue

                cell = ws.cell(row_idx, col_idx)
                value = str(cell.value or "")
                value = value.replace("<", "&lt;").replace(">", "&gt;")

                tag = "th" if row_idx == 1 else "td"
                attrs = ""

                if (row_idx, col_idx) in merge_map:
                    rowspan, colspan = merge_map[(row_idx, col_idx)]
                    if rowspan > 1:
                        attrs += f' rowspan="{rowspan}"'
                    if colspan > 1:
                        attrs += f' colspan="{colspan}"'

                html_parts.append(f"    <{tag}{attrs}>{value}</{tag}>")
            html_parts.append("  </tr>")

        html_parts.append("</table>")
        return "\n".join(html_parts)

    def export(self, output_path: str):
        """Export all tables to OmniDocBench JSON format."""
        data = {
            "dataset_name": self.dataset_name,
            "version": self.version,
            "num_tables": len(self.tables),
            "tables": [t.to_dict() for t in self.tables],
        }

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(data, f, indent=2, ensure_ascii=False)

        logger.info(f"Exported {len(self.tables)} tables to {output_path}")
        return data

    def export_for_teds_eval(
        self,
        predictions_dir: str,
        output_path: str,
    ) -> Dict[str, Any]:
        """
        Export paired ground truth + predictions for TEDS evaluation.

        Expects prediction HTML files in predictions_dir named:
            {table_id}.html

        Output format:
            {
                "pairs": [
                    {"id": "...", "gt_html": "...", "pred_html": "...", "attributes": {...}}
                ]
            }
        """
        pairs = []
        predictions_path = Path(predictions_dir)

        for table in self.tables:
            pred_file = predictions_path / f"{table.id}.html"
            if pred_file.exists():
                pred_html = pred_file.read_text(encoding="utf-8")
                pairs.append({
                    "id": table.id,
                    "gt_html": table.html,
                    "pred_html": pred_html,
                    "attributes": table.attributes.to_stratification_dict(),
                })
            else:
                logger.warning(f"No prediction found for {table.id}")

        output = {
            "dataset_name": self.dataset_name,
            "num_pairs": len(pairs),
            "pairs": pairs,
        }

        with open(output_path, "w", encoding="utf-8") as f:
            json.dump(output, f, indent=2, ensure_ascii=False)

        logger.info(f"Exported {len(pairs)} pairs for TEDS evaluation")
        return output

    def summary(self) -> Dict[str, Any]:
        """Generate a summary of the exported dataset."""
        if not self.tables:
            return {"num_tables": 0}

        total_cells = sum(t.attributes.num_rows * t.attributes.num_cols for t in self.tables)
        merged = sum(1 for t in self.tables if t.attributes.has_merged_cells)
        footnoted = sum(1 for t in self.tables if t.attributes.has_footnotes)
        conditional = sum(1 for t in self.tables if t.attributes.has_conditional_markers)

        return {
            "num_tables": len(self.tables),
            "total_cells": total_cells,
            "avg_rows": sum(t.attributes.num_rows for t in self.tables) / len(self.tables),
            "avg_cols": sum(t.attributes.num_cols for t in self.tables) / len(self.tables),
            "tables_with_merged_cells": merged,
            "tables_with_footnotes": footnoted,
            "tables_with_conditional_markers": conditional,
            "sources": list(set(t.source for t in self.tables)),
        }


# ── CLI usage ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    import argparse

    parser = argparse.ArgumentParser(description="OmniDocBench Exporter")
    parser.add_argument("--xlsx", nargs="+", help="XLSX annotation files")
    parser.add_argument("--source", default="unknown.pdf", help="Source PDF name")
    parser.add_argument("--output", default="omnidocbench_eval.json")
    parser.add_argument("--predictions-dir", help="Directory with prediction HTML files")
    args = parser.parse_args()

    exporter = OmniDocBenchExporter()

    if args.xlsx:
        for xlsx_file in args.xlsx:
            exporter.add_table_from_xlsx(xlsx_file, args.source)

    data = exporter.export(args.output)
    print(f"Exported {data['num_tables']} tables to {args.output}")

    summary = exporter.summary()
    print(f"\nSummary:")
    for k, v in summary.items():
        print(f"  {k}: {v}")

    if args.predictions_dir:
        pairs_output = args.output.replace(".json", "_pairs.json")
        exporter.export_for_teds_eval(args.predictions_dir, pairs_output)
